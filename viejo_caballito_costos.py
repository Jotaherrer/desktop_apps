"""
App de escritorio - Costos
Comando para convertir en ejecutable: pyinstaller --windowed --onedir --icon=./images.ico viejo_caballito.py
"""
import ttkthemes as themes
import tkinter as tk
from tkinter import messagebox, ttk
import os, time
import xlwings as xw
from openpyxl import Workbook, load_workbook


def comprobar_archivo():
    existe = os.path.exists('Costos.xlsx')
    if existe:
        wb = load_workbook(filename='Costos.xlsx')
        ws_costos = wb['Dato Costos']
        ws_obs = wb['Observaciones']
        ws_uni = wb['Unidades']
        print('Apertura exitosa del archivo.')
    else:
        wb = Workbook()
        ws_costos = wb.create_sheet('Dato Costos',0)
        ws_obs = wb.create_sheet('Observaciones',1)
        ws_uni = wb.create_sheet('Unidades',2)
        titulo_costos = ('Hora transacción','Queso','Leche','Pollo','Carne P.','Tapa','Cebolla','Pan','Tomate','Lechuga','Yogur','Agua','Nalga','Empleados','Acelga','Huevos','Servilletas','Yerba','Cafe','Jamón','Puerro','Berenjenas','Papa','Calabaza','Alquiler','Luz','AYSA','Telefono','ABL','Diario','Fumig.','Deterg.','Monotr.','Otros','Otros','Otros','Tarjeta?','Total Fijos y Variables')
        ws_costos.append(titulo_costos)
        titulos_obs = ('Hora transacción', 'Alquiler','Luz','AYSA','Telefono','ABL','Diario','Fumig.','Deterg.','Monotr.','Otros','Otros','Otros','Tarjeta?', 'Total Costos Fijos')
        ws_obs.append(titulos_obs)
        titulos_uni =('Hora transacción', 'Cantidad Queso', 'Costo Queso', 'Cantidad leche', 'Costo leche', 'Cantidad pollo', 'Costo pollo', 'Cantidad carne picada', 'Costo carne picada', 'Cantidad tapa', 'Costo tapa', 'Cantidad cebolla', 'Costo cebolla', 'Cantidad pan', 'Costo pan', 'Cantidad tomate', 'Costo tomate', 'Cantidad lechuga', 'Costo lechuga', 'Cantidad yogur', 'Costo yogur', 'Cantidad agua', 'Costo agua', 'Cantidad nalga', 'Costo nalga', 'Cantidad empleados', 'Costo empleados', 'Cantidad acelga', 'Costo acelga',  'Cantidad huevos', 'Costo huevos', 'Cantidad servilletas', 'Costo servilletas', 'Cantidad yerba', 'Costo yerba', 'Cantidad cafe', 'Costo cafe', 'Cantidad jamon', 'Costo jamon', 'Cantidad puerro', 'Costo puerro', 'Cantidad berenjena', 'Costo berenjena', 'Cantidad papa', 'Costo papa', 'Cantidad calabaza', 'Costo calabaza',)
        ws_uni.append(titulos_uni)
        wb.save(filename='Costos.xlsx')
        print('Creación exitosa del archivo')


def guardar_datos_costos(pedido):
    wb = load_workbook(filename='Costos.xlsx')
    wb['Dato Costos'].append(pedido)
    wb.save('Costos.xlsx')
    print("Carga exitosa del costo!!")


def guardar_datos_obs(info):
    wb = load_workbook(filename='Costos.xlsx')
    wb['Observaciones'].append(info)
    wb.save('Costos.xlsx')
    print("Carga exitosa de las observaciones!!")


def guardar_datos_unidades(info):
    wb = load_workbook(filename='Costos.xlsx')
    wb['Unidades'].append(info)
    wb.save('Costos.xlsx')
    print("Carga exitosa de las unidades!!")


def cambiar_tarjeta_valor():
    if tarjeta_valor.get() == int(1):
        tarjeta_valor.set(0)
        print('Borrado boton tarjeta_valor')
    else:
        print('Boton no tildado..')
        print(tarjeta_valor.get())
        print(type(tarjeta_valor.get()))
        pass


def contenido(caja_precio, caja_cantidad, total):
    try:
        cantidad = int(caja_cantidad.get())
        precio = int(caja_precio.get())
        variable = int(caja_precio.get()) * int(caja_cantidad.get())
        total.insert('0',variable)
    except:
        if (caja_precio.get() == '') | (caja_cantidad.get() == ''):
            variable = 0
            cantidad = 0
            precio = 0
        else:
            messagebox.showinfo(title='Error', message='Ingresar un número válido')
    return variable, cantidad, precio


def contenido_fijos(caja_precio, caja_obs):
    try:
        variable = int(caja_precio.get())
        observacion = caja_obs.get()
    except:
        if caja_precio.get() == '':
            variable = 0
            observacion = ''
        else:
            messagebox.showinfo(title='Error', message='Ingresar un número válido')
    return variable, observacion


def mult():
    hora = time.asctime()
    if caja_total.get() == '':
        ## COMPLETA VALORES CON FUNCION 'CONTENIDO'
        # COSTOS VARIABLES
        queso, q_queso, p_queso = contenido(caja_p_queso, caja_q_queso, caja_t_queso)
        leche, q_leche, p_leche = contenido(caja_p_leche, caja_q_leche, caja_t_leche)
        pollo, q_pollo, p_pollo = contenido(caja_p_pollo, caja_q_pollo, caja_t_pollo)
        carne_p, q_carne_p, p_carne_p = contenido(caja_p_carnep, caja_q_carnep, caja_t_carnep)
        tapa, q_tapa, p_tapa = contenido(caja_p_tapa, caja_q_tapa, caja_t_tapa)
        cebolla, q_cebolla, p_cebolla = contenido(caja_p_cebolla, caja_q_cebolla, caja_t_cebolla)
        pan, q_pan, p_pan = contenido(caja_p_pan, caja_q_pan, caja_t_pan)
        tomate, q_tomate, p_tomate = contenido(caja_p_tomate, caja_q_tomate, caja_t_tomate)
        lechuga, q_lechuga, p_lechuga = contenido(caja_p_lechuga, caja_q_lechuga, caja_t_lechuga)
        yogur, q_yogur, p_yogur = contenido(caja_p_yogur, caja_q_yogur, caja_t_yogur)
        agua, q_agua, p_agua = contenido(caja_p_agua, caja_q_agua, caja_t_agua)
        nalga, q_nalga, p_nalga = contenido(caja_p_nalga, caja_q_nalga, caja_t_nalga)
        empleados, q_empleados, p_empleados = contenido(caja_p_empleados, caja_q_empleados, caja_t_empleados)
        acelga, q_acelga, p_acelga = contenido(caja_p_acelga, caja_q_acelga, caja_t_acelga)
        huevos, q_huevos, p_huevos = contenido(caja_p_huevos, caja_q_huevos, caja_t_huevos)
        servilletas, q_servilletas, p_servilletas = contenido(caja_p_servilletas, caja_q_servilletas, caja_t_servilletas)
        yerba, q_yerba, p_yerba = contenido(caja_p_yerba, caja_q_yerba, caja_t_yerba)
        cafe, q_cafe, p_cafe = contenido(caja_p_cafe, caja_q_cafe, caja_t_cafe)
        jamon, q_jamon, p_jamon = contenido(caja_p_jamon, caja_q_jamon, caja_t_jamon)
        puerro, q_puerro, p_puerro = contenido(caja_p_puerro, caja_q_puerro, caja_t_puerro)
        beren, q_beren, p_beren = contenido(caja_p_beren, caja_q_beren, caja_t_beren)
        papa, q_papa, p_papa = contenido(caja_p_papa, caja_q_papa, caja_t_papa)
        cala, q_cala, p_cala = contenido(caja_p_calabaza, caja_q_calabaza, caja_t_calabaza)

        # COSTOS FIJOS
        alquiler = contenido_fijos(caja_p_alquiler,caja_obs_alquiler)[0]
        luz = contenido_fijos(caja_p_luz, caja_obs_luz)[0]
        agua_servicio = contenido_fijos(caja_p_agua_servicio, caja_obs_agua)[0]
        telefono = contenido_fijos(caja_p_telefono, caja_obs_telefono)[0]
        abl = contenido_fijos(caja_p_abl, caja_obs_abl)[0]
        diario = contenido_fijos(caja_p_diario, caja_obs_diario)[0]
        fumigador = contenido_fijos(caja_p_fumigador, caja_obs_fumigador)[0]
        detergente = contenido_fijos(caja_p_detergente, caja_obs_fumigador)[0]
        monotributo = contenido_fijos(caja_p_monotributo, caja_obs_monotributo)[0]
        otros1 = contenido_fijos(caja_p_otros1, caja_obs_otros1)[0]
        otros2 = contenido_fijos(caja_p_otros2, caja_obs_otros2)[0]
        otros3 = contenido_fijos(caja_p_otros3, caja_obs_otros3)[0]

    else:
        # BORRA CASILLEROS COMPLETOS
        for caja in cajas:
            try:
                caja.delete(0, tk.END)
            except:
                pass

        cambiar_tarjeta_valor()

        # COMPLETA NUEVAMENTE VALORES
        # COSTOS VARIABLES
        queso, q_queso, p_queso = contenido(caja_p_queso, caja_q_queso, caja_t_queso)
        leche, q_leche, p_leche = contenido(caja_p_leche, caja_q_leche, caja_t_leche)
        pollo, q_pollo, p_pollo = contenido(caja_p_pollo, caja_q_pollo, caja_t_pollo)
        carne_p, q_carne_p, p_carne_p = contenido(caja_p_carnep, caja_q_carnep, caja_t_carnep)
        tapa, q_tapa, p_tapa = contenido(caja_p_tapa, caja_q_tapa, caja_t_tapa)
        cebolla, q_cebolla, p_cebolla = contenido(caja_p_cebolla, caja_q_cebolla, caja_t_cebolla)
        pan, q_pan, p_pan = contenido(caja_p_pan, caja_q_pan, caja_t_pan)
        tomate, q_tomate, p_tomate = contenido(caja_p_tomate, caja_q_tomate, caja_t_tomate)
        lechuga, q_lechuga, p_lechuga = contenido(caja_p_lechuga, caja_q_lechuga, caja_t_lechuga)
        yogur, q_yogur, p_yogur = contenido(caja_p_yogur, caja_q_yogur, caja_t_yogur)
        agua, q_agua, p_agua = contenido(caja_p_agua, caja_q_agua, caja_t_agua)
        nalga, q_nalga, p_nalga = contenido(caja_p_nalga, caja_q_nalga, caja_t_nalga)
        empleados, q_empleados, p_empleados = contenido(caja_p_empleados, caja_q_empleados, caja_t_empleados)
        acelga, q_acelga, p_acelga = contenido(caja_p_acelga, caja_q_acelga, caja_t_acelga)
        huevos, q_huevos, p_huevos = contenido(caja_p_huevos, caja_q_huevos, caja_t_huevos)
        servilletas, q_servilletas, p_servilletas = contenido(caja_p_servilletas, caja_q_servilletas, caja_t_servilletas)
        yerba, q_yerba, p_yerba = contenido(caja_p_yerba, caja_q_yerba, caja_t_yerba)
        cafe, q_cafe, p_cafe = contenido(caja_p_cafe, caja_q_cafe, caja_t_cafe)
        jamon, q_jamon, p_jamon = contenido(caja_p_jamon, caja_q_jamon, caja_t_jamon)
        puerro, q_puerro, p_puerro = contenido(caja_p_puerro, caja_q_puerro, caja_t_puerro)
        beren, q_beren, p_beren = contenido(caja_p_beren, caja_q_beren, caja_t_beren)
        papa, q_papa, p_papa = contenido(caja_p_papa, caja_q_papa, caja_t_papa)
        cala, q_cala, p_cala = contenido(caja_p_calabaza, caja_q_calabaza, caja_t_calabaza)

        # COSTOS FIJOS
        alquiler = contenido_fijos(caja_p_alquiler,caja_obs_alquiler)[0]
        luz = contenido_fijos(caja_p_luz, caja_obs_luz)[0]
        agua_servicio = contenido_fijos(caja_p_agua_servicio, caja_obs_agua)[0]
        telefono = contenido_fijos(caja_p_telefono, caja_obs_telefono)[0]
        abl = contenido_fijos(caja_p_abl, caja_obs_abl)[0]
        diario = contenido_fijos(caja_p_diario, caja_obs_diario)[0]
        fumigador = contenido_fijos(caja_p_fumigador, caja_obs_fumigador)[0]
        detergente = contenido_fijos(caja_p_detergente, caja_obs_fumigador)[0]
        monotributo = contenido_fijos(caja_p_monotributo, caja_obs_monotributo)[0]
        otros1 = contenido_fijos(caja_p_otros1, caja_obs_otros1)[0]
        otros2 = contenido_fijos(caja_p_otros2, caja_obs_otros2)[0]
        otros3 = contenido_fijos(caja_p_otros3, caja_obs_otros3)[0]

    pago_tarjeta = checkbox_clicked()
    costos_varios = queso + leche + pollo + carne_p + tapa + cebolla + pan + tomate + lechuga+yogur+agua+nalga+empleados + acelga + huevos + servilletas + yerba + cafe + jamon + puerro + beren + papa + cala
    costos_fijos = alquiler + luz + agua_servicio + telefono + abl + diario + fumigador + detergente + monotributo + otros1 + otros2 + otros3
    facturacion = costos_varios + costos_fijos
    caja_total.insert('0',facturacion)

    # PASAJE A EXCEL
    al_excel = [hora, queso,leche,pollo,carne_p,tapa,cebolla,pan,tomate,lechuga,yogur, agua,nalga,empleados,acelga,huevos,servilletas,yerba,cafe,
                jamon,puerro,beren,papa,cala,alquiler,luz,agua_servicio,telefono,abl,diario,fumigador,detergente,monotributo,otros1, otros2, otros3,pago_tarjeta,facturacion]
    al_excel_obs = [hora, contenido_fijos(caja_p_alquiler,caja_obs_alquiler)[1], contenido_fijos(caja_p_luz, caja_obs_luz)[1],
                    contenido_fijos(caja_p_agua_servicio, caja_obs_agua)[1], contenido_fijos(caja_p_telefono, caja_obs_telefono)[1],
                    contenido_fijos(caja_p_abl, caja_obs_abl)[1], contenido_fijos(caja_p_diario, caja_obs_diario)[1],
                    contenido_fijos(caja_p_fumigador, caja_obs_fumigador)[1], contenido_fijos(caja_p_detergente, caja_obs_detergente)[1],
                    contenido_fijos(caja_p_monotributo, caja_obs_monotributo)[1],contenido_fijos(caja_p_otros1, caja_obs_otros1)[1],
                    contenido_fijos(caja_p_otros2, caja_obs_otros2)[1],contenido_fijos(caja_p_otros3, caja_obs_otros3)[1],pago_tarjeta,costos_fijos]
    al_excel_unidades = [hora, q_queso, p_queso, q_leche, p_leche, q_pollo, p_pollo, q_carne_p, p_carne_p, q_tapa, p_tapa, q_cebolla, p_cebolla,
                         q_pan, p_pan, q_tomate, p_tomate, q_lechuga, p_lechuga, q_yogur, p_yogur, q_agua, p_agua, q_nalga, p_nalga,
                         q_empleados, p_empleados, q_acelga, p_acelga, q_huevos, p_huevos, q_servilletas, p_servilletas, q_yerba, p_yerba,
                         q_cafe, p_cafe, q_jamon, p_jamon, q_puerro, p_puerro, q_beren, p_beren, q_papa, p_papa, q_cala, p_cala]
    guardar_datos_costos(al_excel)
    guardar_datos_obs(al_excel_obs)
    guardar_datos_unidades(al_excel_unidades)


def confirmar():
    if caja_total == '':
        messagebox.showinfo(title='Error', message='Ingresar datos numéricos en el registro.')
    else:
        for caja in cajas:
            caja.delete(0, tk.END)

        caja_total.delete(0,tk.END)
        cambiar_tarjeta_valor()


def borrar_datos():
    for caja in cajas:
        caja.delete(0, tk.END)
    cambiar_tarjeta_valor()


def checkbox_clicked():
    rta = tarjeta_valor.get()
    return rta


### EXCEL INICIAL
comprobar_archivo()

### APP DE ESCRITORIO
ventana = themes.ThemedTk()
ventana.set_theme('plastik') # Other 'ventana.get_themes()'
ventana.config(height=750, width=900)
ventana.title("Aplicación de costos - Viejo Caballito Bar")

### ETIQUETAS
cantidad = ttk.Label(text='Cantidad (Kg./Lit.)').place(x=115,y=40)
precio = ttk.Label(text='Precio por unidad').place(x=230,y=40)
total = ttk.Label(text='Costo total').place(x=355,y=40)
precio_fijos = ttk.Label(text='Costo total').place(x=590, y=40)
obser_fijos = ttk.Label(text='Observación a realizar').place(x=680, y=40)

cv = ttk.Label(text='Panel de costos VARIOS:').place(x=20,y=15)
cf = ttk.Label(text='Panel de costos FIJOS:').place(x=500,y=15)

### CHECKBOX
tarjeta_valor = tk.IntVar()
tarjeta = ttk.Checkbutton(text='Pago con tarjeta?', variable=tarjeta_valor, command=checkbox_clicked)
tarjeta.place(x=500, y=600)


## COSTOS VARIOS
ttk.Label(text="COMIDAS").place(x=20, y=60)
label_queso = ttk.Label(text='Horma queso  ==> ')
label_queso.place(x=20,y=75)
label_leche = ttk.Label(text='Leche  ==> ')
label_leche.place(x=20,y=100)
label_pollo = ttk.Label(text='Pollo  ==> ')
label_pollo.place(x=20,y=125)
label_carne_picada = ttk.Label(text='Carne Picada  ==> ')
label_carne_picada.place(x=20,y=150)
label_tapa = ttk.Label(text='Tapas  ==> ')
label_tapa.place(x=20,y=175)
label_cebolla = ttk.Label(text='Cebolla  ==> ')
label_cebolla.place(x=20,y=200)
label_pan = ttk.Label(text='Pan  ==> ')
label_pan.place(x=20,y=225)
label_tomate = ttk.Label(text='Tomate  ==> ')
label_tomate.place(x=20,y=250)
label_lechuga = ttk.Label(text='Lechuga  ==> ')
label_lechuga.place(x=20,y=275)
label_yogur = ttk.Label(text='Yogur  ==> ')
label_yogur.place(x=20,y=300)
label_agua = ttk.Label(text='Agua  ==> ')
label_agua.place(x=20,y=325)
label_nalga = ttk.Label(text='Nalga  ==> ')
label_nalga.place(x=20,y=350)
label_empleada = ttk.Label(text='Empleados  ==> ')
label_empleada.place(x=20,y=375)
label_acelga = ttk.Label(text='Acelga  ==> ')
label_acelga.place(x=20,y=400)
label_huevos = ttk.Label(text='Huevos  ==> ')
label_huevos.place(x=20,y=425)
label_servilletas = ttk.Label(text='Servilletas  ==> ')
label_servilletas.place(x=20,y=450)
label_yerba = ttk.Label(text='Yerba  ==> ')
label_yerba.place(x=20,y=475)
label_cafe = ttk.Label(text='Café  ==> ')
label_cafe.place(x=20,y=500)
label_jamon = ttk.Label(text='Jamón  ==> ')
label_jamon.place(x=20,y=525)
label_puerro = ttk.Label(text='Puerro  ==> ')
label_puerro.place(x=20,y=550)
label_beren = ttk.Label(text='Berenjenas  ==> ')
label_beren.place(x=20,y=575)
label_papa = ttk.Label(text='Papas  ==> ')
label_papa.place(x=20,y=600)
label_cala = ttk.Label(text='Calabaza  ==> ')
label_cala.place(x=20,y=625)

## COSTOS FIJOS
label_alquiler = ttk.Label(text='Alquiler  ==> ')
label_alquiler.place(x=500,y=60)
label_luz = ttk.Label(text='Luz  ==> ')
label_luz.place(x=500,y=85)
label_agua_servicio = ttk.Label(text='AYSA ==> ')
label_agua_servicio.place(x=500,y=110)
label_telefono = ttk.Label(text='Teléfono  ==> ')
label_telefono.place(x=500,y=135)
label_abl = ttk.Label(text='ABL  ==> ')
label_abl.place(x=500,y=160)
label_diario = ttk.Label(text='Diario  ==> ')
label_diario.place(x=500,y=185)
label_fumigador = ttk.Label(text='Fumigador  ==> ')
label_fumigador.place(x=500,y=210)
label_detergente = ttk.Label(text='Detergente  ==> ')
label_detergente.place(x=500,y=235)
label_monotributo = ttk.Label(text='Monotributo  => ')
label_monotributo.place(x=500,y=260)
label_otros1 = ttk.Label(text='Otros  => ')
label_otros1.place(x=500,y=285)
label_otros2 = ttk.Label(text='Otros  => ')
label_otros2.place(x=500,y=310)
label_otros3 = ttk.Label(text='Otros  => ')
label_otros3.place(x=500,y=335)

### CAJAS
## COSTOS VARIOS
caja_q_queso = ttk.Entry()
caja_q_queso.place(x=125, y=75,width=80)
caja_q_queso.insert(tk.END,'')
caja_p_queso = ttk.Entry()
caja_p_queso.place(x=235, y=75,width=80)
caja_p_queso.insert(tk.END,'')
caja_t_queso = ttk.Entry()
caja_t_queso.place(x=345, y=75,width=80)
caja_t_queso.insert(tk.END,'')

caja_q_leche = ttk.Entry()
caja_q_leche.place(x=125, y=100,width=80)
caja_q_leche.insert(tk.END,'')
caja_p_leche = ttk.Entry()
caja_p_leche.place(x=235, y=100,width=80)
caja_p_leche.insert(tk.END,'')
caja_t_leche = ttk.Entry()
caja_t_leche.place(x=345, y=100,width=80)
caja_t_leche.insert(tk.END,'')

caja_q_pollo = ttk.Entry()
caja_q_pollo.place(x=125, y=125,width=80)
caja_q_pollo.insert(tk.END,'')
caja_p_pollo = ttk.Entry()
caja_p_pollo.place(x=235, y=125,width=80)
caja_p_pollo.insert(tk.END,'')
caja_t_pollo = ttk.Entry()
caja_t_pollo.place(x=345, y=125,width=80)
caja_t_pollo.insert(tk.END,'')

caja_q_carnep = ttk.Entry()
caja_q_carnep.place(x=125, y=150,width=80)
caja_q_carnep.insert(tk.END,'')
caja_p_carnep = ttk.Entry()
caja_p_carnep.place(x=235, y=150,width=80)
caja_p_carnep.insert(tk.END,'')
caja_t_carnep = ttk.Entry()
caja_t_carnep.place(x=345, y=150,width=80)
caja_t_carnep.insert(tk.END,'')

caja_q_tapa = ttk.Entry()
caja_q_tapa.place(x=125, y=175,width=80)
caja_q_tapa.insert(tk.END,'')
caja_p_tapa = ttk.Entry()
caja_p_tapa.place(x=235, y=175,width=80)
caja_p_tapa.insert(tk.END,'')
caja_t_tapa = ttk.Entry()
caja_t_tapa.place(x=345, y=175,width=80)
caja_t_tapa.insert(tk.END,'')

caja_q_cebolla = ttk.Entry()
caja_q_cebolla.place(x=125, y=200,width=80)
caja_q_cebolla.insert(tk.END,'')
caja_p_cebolla = ttk.Entry()
caja_p_cebolla.place(x=235, y=200,width=80)
caja_p_cebolla.insert(tk.END,'')
caja_t_cebolla = ttk.Entry()
caja_t_cebolla.place(x=345, y=200,width=80)
caja_t_cebolla.insert(tk.END,'')

caja_q_pan = ttk.Entry()
caja_q_pan.place(x=125, y=225,width=80)
caja_q_pan.insert(tk.END,'')
caja_p_pan = ttk.Entry()
caja_p_pan.place(x=235, y=225,width=80)
caja_p_pan.insert(tk.END,'')
caja_t_pan = ttk.Entry()
caja_t_pan.place(x=345, y=225,width=80)
caja_t_pan.insert(tk.END,'')

caja_q_tomate = ttk.Entry()
caja_q_tomate.place(x=125, y=250,width=80)
caja_q_tomate.insert(tk.END,'')
caja_p_tomate = ttk.Entry()
caja_p_tomate.place(x=235, y=250,width=80)
caja_p_tomate.insert(tk.END,'')
caja_t_tomate = ttk.Entry()
caja_t_tomate.place(x=345, y=250,width=80)
caja_t_tomate.insert(tk.END,'')

caja_q_lechuga = ttk.Entry()
caja_q_lechuga.place(x=125, y=275,width=80)
caja_q_lechuga.insert(tk.END,'')
caja_p_lechuga = ttk.Entry()
caja_p_lechuga.place(x=235, y=275,width=80)
caja_p_lechuga.insert(tk.END,'')
caja_t_lechuga = ttk.Entry()
caja_t_lechuga.place(x=345, y=275,width=80)
caja_t_lechuga.insert(tk.END,'')

caja_q_yogur = ttk.Entry()
caja_q_yogur.place(x=125, y=300,width=80)
caja_q_yogur.insert(tk.END,'')
caja_p_yogur = ttk.Entry()
caja_p_yogur.place(x=235, y=300,width=80)
caja_p_yogur.insert(tk.END,'')
caja_t_yogur = ttk.Entry()
caja_t_yogur.place(x=345, y=300,width=80)
caja_t_yogur.insert(tk.END,'')

caja_q_agua = ttk.Entry()
caja_q_agua.place(x=125, y=325,width=80)
caja_q_agua.insert(tk.END,'')
caja_p_agua = ttk.Entry()
caja_p_agua.place(x=235, y=325,width=80)
caja_p_agua.insert(tk.END,'')
caja_t_agua = ttk.Entry()
caja_t_agua.place(x=345, y=325,width=80)
caja_t_agua.insert(tk.END,'')

caja_q_nalga = ttk.Entry()
caja_q_nalga.place(x=125, y=350,width=80)
caja_q_nalga.insert(tk.END,'')
caja_p_nalga = ttk.Entry()
caja_p_nalga.place(x=235, y=350,width=80)
caja_p_nalga.insert(tk.END,'')
caja_t_nalga = ttk.Entry()
caja_t_nalga.place(x=345, y=350,width=80)
caja_t_nalga.insert(tk.END,'')

caja_q_empleados = ttk.Entry()
caja_q_empleados.place(x=125, y=375,width=80)
caja_q_empleados.insert(tk.END,'')
caja_p_empleados = ttk.Entry()
caja_p_empleados.place(x=235, y=375,width=80)
caja_p_empleados.insert(tk.END,'')
caja_t_empleados = ttk.Entry()
caja_t_empleados.place(x=345, y=375,width=80)
caja_t_empleados.insert(tk.END,'')

caja_q_acelga = ttk.Entry()
caja_q_acelga.place(x=125, y=400,width=80)
caja_q_acelga.insert(tk.END,'')
caja_p_acelga = ttk.Entry()
caja_p_acelga.place(x=235, y=400,width=80)
caja_p_acelga.insert(tk.END,'')
caja_t_acelga = ttk.Entry()
caja_t_acelga.place(x=345, y=400,width=80)
caja_t_acelga.insert(tk.END,'')

caja_q_huevos = ttk.Entry()
caja_q_huevos.place(x=125, y=425,width=80)
caja_q_huevos.insert(tk.END,'')
caja_p_huevos = ttk.Entry()
caja_p_huevos.place(x=235, y=425,width=80)
caja_p_huevos.insert(tk.END,'')
caja_t_huevos = ttk.Entry()
caja_t_huevos.place(x=345, y=425,width=80)
caja_t_huevos.insert(tk.END,'')

caja_q_servilletas = ttk.Entry()
caja_q_servilletas.place(x=125, y=450,width=80)
caja_q_servilletas.insert(tk.END,'')
caja_p_servilletas = ttk.Entry()
caja_p_servilletas.place(x=235, y=450,width=80)
caja_p_servilletas.insert(tk.END,'')
caja_t_servilletas = ttk.Entry()
caja_t_servilletas.place(x=345, y=450,width=80)
caja_t_servilletas.insert(tk.END,'')

caja_q_yerba = ttk.Entry()
caja_q_yerba.place(x=125, y=475,width=80)
caja_q_yerba.insert(tk.END,'')
caja_p_yerba = ttk.Entry()
caja_p_yerba.place(x=235, y=475,width=80)
caja_p_yerba.insert(tk.END,'')
caja_t_yerba = ttk.Entry()
caja_t_yerba.place(x=345, y=475,width=80)
caja_t_yerba.insert(tk.END,'')

caja_q_cafe = ttk.Entry()
caja_q_cafe.place(x=125, y=500,width=80)
caja_q_cafe.insert(tk.END,'')
caja_p_cafe = ttk.Entry()
caja_p_cafe.place(x=235, y=500,width=80)
caja_p_cafe.insert(tk.END,'')
caja_t_cafe = ttk.Entry()
caja_t_cafe.place(x=345, y=500,width=80)
caja_t_cafe.insert(tk.END,'')

caja_q_jamon = ttk.Entry()
caja_q_jamon.place(x=125, y=525,width=80)
caja_q_jamon.insert(tk.END,'')
caja_p_jamon = ttk.Entry()
caja_p_jamon.place(x=235, y=525,width=80)
caja_p_jamon.insert(tk.END,'')
caja_t_jamon = ttk.Entry()
caja_t_jamon.place(x=345, y=525,width=80)
caja_t_jamon.insert(tk.END,'')

caja_q_puerro = ttk.Entry()
caja_q_puerro.place(x=125, y=550,width=80)
caja_q_puerro.insert(tk.END,'')
caja_p_puerro = ttk.Entry()
caja_p_puerro.place(x=235, y=550,width=80)
caja_p_puerro.insert(tk.END,'')
caja_t_puerro = ttk.Entry()
caja_t_puerro.place(x=345, y=550,width=80)
caja_t_puerro.insert(tk.END,'')

caja_q_beren = ttk.Entry()
caja_q_beren.place(x=125, y=575,width=80)
caja_q_beren.insert(tk.END,'')
caja_p_beren = ttk.Entry()
caja_p_beren.place(x=235, y=575,width=80)
caja_p_beren.insert(tk.END,'')
caja_t_beren = ttk.Entry()
caja_t_beren.place(x=345, y=575,width=80)
caja_t_beren.insert(tk.END,'')

caja_q_papa = ttk.Entry()
caja_q_papa.place(x=125, y=600,width=80)
caja_q_papa.insert(tk.END,'')
caja_p_papa = ttk.Entry()
caja_p_papa.place(x=235, y=600,width=80)
caja_p_papa.insert(tk.END,'')
caja_t_papa = ttk.Entry()
caja_t_papa.place(x=345, y=600,width=80)
caja_t_papa.insert(tk.END,'')

caja_q_calabaza = ttk.Entry()
caja_q_calabaza.place(x=125, y=625,width=80)
caja_q_calabaza.insert(tk.END,'')
caja_p_calabaza = ttk.Entry()
caja_p_calabaza.place(x=235, y=625,width=80)
caja_p_calabaza.insert(tk.END,'')
caja_t_calabaza = ttk.Entry()
caja_t_calabaza.place(x=345, y=625,width=80)
caja_t_calabaza.insert(tk.END,'')

caja_total = ttk.Entry()
caja_total.place(x=500, y=625, width=160)
caja_total.insert(tk.END,'')

## COSTOS FIJOS
caja_p_alquiler = ttk.Entry()
caja_p_alquiler.place(x=590, y=60,width=80)
caja_p_alquiler.insert(tk.END,'')
caja_obs_alquiler = ttk.Entry()
caja_obs_alquiler.place(x=680, y=60,width=180)
caja_obs_alquiler.insert(tk.END,'')

caja_p_luz = ttk.Entry()
caja_p_luz.place(x=590, y=85,width=80)
caja_p_luz.insert(tk.END,'')
caja_obs_luz = ttk.Entry()
caja_obs_luz.place(x=680, y=85,width=180)
caja_obs_luz.insert(tk.END,'')

caja_p_agua_servicio = ttk.Entry()
caja_p_agua_servicio.place(x=590, y=110,width=80)
caja_p_agua_servicio.insert(tk.END,'')
caja_obs_agua = ttk.Entry()
caja_obs_agua.place(x=680, y=110,width=180)
caja_obs_agua.insert(tk.END,'')

caja_p_telefono = ttk.Entry()
caja_p_telefono.place(x=590, y=135,width=80)
caja_p_telefono.insert(tk.END,'')
caja_obs_telefono = ttk.Entry()
caja_obs_telefono.place(x=680, y=135,width=180)
caja_obs_telefono.insert(tk.END,'')

caja_p_abl = ttk.Entry()
caja_p_abl.place(x=590, y=160,width=80)
caja_p_abl.insert(tk.END,'')
caja_obs_abl = ttk.Entry()
caja_obs_abl.place(x=680, y=160,width=180)
caja_obs_abl.insert(tk.END,'')

caja_p_diario = ttk.Entry()
caja_p_diario.place(x=590, y=185,width=80)
caja_p_diario.insert(tk.END,'')
caja_obs_diario = ttk.Entry()
caja_obs_diario.place(x=680, y=185,width=180)
caja_obs_diario.insert(tk.END,'')

caja_p_fumigador = ttk.Entry()
caja_p_fumigador.place(x=590, y=210,width=80)
caja_p_fumigador.insert(tk.END,'')
caja_obs_fumigador = ttk.Entry()
caja_obs_fumigador.place(x=680, y=210,width=180)
caja_obs_fumigador.insert(tk.END,'')

caja_p_detergente = ttk.Entry()
caja_p_detergente.place(x=590, y=235,width=80)
caja_p_detergente.insert(tk.END,'')
caja_obs_detergente = ttk.Entry()
caja_obs_detergente.place(x=680, y=235,width=180)
caja_obs_detergente.insert(tk.END,'')

caja_p_monotributo = ttk.Entry()
caja_p_monotributo.place(x=590, y=260,width=80)
caja_p_monotributo.insert(tk.END,'')
caja_obs_monotributo = ttk.Entry()
caja_obs_monotributo.place(x=680, y=260,width=180)
caja_obs_monotributo.insert(tk.END,'')

caja_p_otros1 = ttk.Entry()
caja_p_otros1.place(x=590, y=285,width=80)
caja_p_otros1.insert(tk.END,'')
caja_obs_otros1 = ttk.Entry()
caja_obs_otros1.place(x=680, y=285,width=180)
caja_obs_otros1.insert(tk.END,'')

caja_p_otros2 = ttk.Entry()
caja_p_otros2.place(x=590, y=310,width=80)
caja_p_otros2.insert(tk.END,'')
caja_obs_otros2 = ttk.Entry()
caja_obs_otros2.place(x=680, y=310,width=180)
caja_obs_otros2.insert(tk.END,'')

caja_p_otros3 = ttk.Entry()
caja_p_otros3.place(x=590, y=335,width=80)
caja_p_otros3.insert(tk.END,'')
caja_obs_otros3 = ttk.Entry()
caja_obs_otros3.place(x=680, y=335,width=180)
caja_obs_otros3.insert(tk.END,'')

cajas = [caja_q_queso, caja_p_queso, caja_t_queso, caja_q_leche, caja_p_leche, caja_t_leche,
         caja_p_pollo, caja_q_pollo, caja_t_pollo, caja_p_carnep, caja_t_carnep,
         caja_q_carnep, caja_p_queso, caja_t_queso, caja_q_tapa, caja_p_tapa, caja_t_tapa,
         caja_q_cebolla, caja_p_cebolla, caja_t_cebolla, caja_p_pan, caja_q_pan, caja_t_pan,
         caja_q_tomate, caja_p_tomate, caja_t_tomate, caja_q_lechuga, caja_p_lechuga, caja_t_lechuga,
         caja_q_yogur, caja_p_yogur, caja_t_yogur, caja_q_agua, caja_p_agua, caja_t_agua,
         caja_q_nalga, caja_p_nalga, caja_t_nalga, caja_q_empleados, caja_p_empleados, caja_t_empleados,
         caja_q_acelga, caja_p_acelga, caja_t_acelga, caja_q_huevos, caja_p_huevos, caja_t_huevos,
         caja_q_servilletas, caja_p_servilletas, caja_q_yerba, caja_p_yerba, caja_q_cafe, caja_p_cafe, caja_t_cafe,
         caja_q_jamon, caja_p_jamon, caja_t_jamon, caja_q_puerro, caja_p_puerro, caja_t_servilletas,
         caja_q_beren, caja_p_beren, caja_t_beren, caja_q_papa, caja_p_papa, caja_t_yerba,
         caja_q_calabaza, caja_p_calabaza, caja_t_calabaza, caja_p_alquiler, caja_obs_alquiler,
         caja_p_luz, caja_obs_luz, caja_p_agua_servicio, caja_obs_agua, caja_t_puerro, caja_t_papa,
         caja_p_telefono, caja_obs_telefono, caja_p_abl, caja_obs_abl, caja_p_diario, caja_obs_diario,
         caja_p_fumigador, caja_obs_fumigador, caja_p_detergente, caja_obs_detergente, caja_p_monotributo,
         caja_obs_monotributo, caja_p_otros1, caja_obs_otros1, caja_p_otros2, caja_obs_otros2,
         caja_p_otros3, caja_obs_otros3]

## BOTONES - MESSAGE BOX
costos = ttk.Button(text='Totalizar costos',command=mult)
costos.place(x=500,y=675)
costos.winfo_class()

confirmacion = ttk.Button(text='Confirmar costos', command=confirmar)
confirmacion.place(x=595,y=675)

boton_borrar = ttk.Button(text='Borrar datos', command=borrar_datos)
boton_borrar.place(x=700,y=675)

ventana.mainloop()

