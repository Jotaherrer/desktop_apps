"""
App de escritorio - Pedidos
Comando para convertir en ejecutable: pyinstaller --windowed --onedir --icon=./images.ico viejo_caballito.py
"""
import ttkthemes as themes
import tkinter as tk
from tkinter import messagebox, ttk
import os, time
import xlwings as xw
from openpyxl import Workbook, load_workbook

lista_venta = []
lista_excel = []
empanadas_de_carne = []
empanadas_de_pollo = []
empanadas_de_jq = []
empanadas_de_verduda = []
empanadas_de_cq = []
lista_tarta_jq = []
lista_tarta_puerro = []
lista_tarta_beren = []
lista_tarta_acelga_queso = []
lista_tarta_acelga_cala = []
lista_tarta_cala = []
lista_tarta_zapa = []
lista_tarta_jqch = []
lista_gaseosa = []
lista_agua = []
lista_cerveza = []
lista_menu = []
lista_menu_sin = []
lista_tortilla = []
lista_ensalada =[]
lista_ensalada_chica =[]
lista_por_papas = []
lista_ome = []
lista_cafe = []
lista_jarrito = []
lista_cafe_leche = []
lista_lagrima = []
lista_te = []
lista_cafe_llevar = []
lista_alfajor = []
lista_medialuna = []
lista_ensa_fruta = []
lista_flan = []

precios = {'empanadas': 60, 'tartas': 200,
           'platos': {'plato_sin_guar': 230, 'plato': 280, 'tortilla': 220, 'ensalada': 250, 'ensa_chica': 150,
                      'porcion_papas': 160, 'omelette': 200},
           'cafeteria': {'cafe_chico': 80, 'jarrito': 90, 'cafe_leche': 120, 'lagrima': 90, 'te': 80,
                         'cafe_llevar': 90, 'alfa': 60,'medialuna': 30},
           'postre': {'ensa_fruta': 150,'flan': 100},
           'bebida': {'gaseosa': 80,'agua': 75,'cerveza': 100}}


def comprobar_archivo():
    existe = os.path.exists('Ventas.xlsx')
    if existe:
        wb = load_workbook(filename='Ventas.xlsx')
        ws = wb.active
        print('Apertura exitosa del archivo.')
    else:
        wb = Workbook()
        ws = wb.active
        titulo = ('Hora transacción',"Emp. Carne", 'Emp. JQ', 'Emp. Pollo', 'Emp. Verdura', 'Emp. CQ', 'Tar. JQ', 'Tar. Puerro',
                  'Tar. Beren.', 'Tar. Acelga y Q.', 'Tar. Acelga y C.','Tar. Calab.', 'Tar. Zapa.','Tar. JQCH', 'Gaseosa', 'Agua',
                  'Cerveza','Plato del Día','Plato S/ Guarn.','Tortilla', 'Ensalada', 'Ensa. chica', 'Papas', 'Omelette', 'Cafe chico',
                  'Jarrito', 'Cafe C/Leche', 'Lagrima', 'Te', 'Cafe P/llevar', 'Alfajores', 'Medialunas', 'Ensa. Fruta', 'Flan',
                  'Descuentos','Tarjeta D.','Total')
        ws.append(titulo)
        wb.save(filename='Ventas.xlsx')
        print('Creación exitosa del archivo')


def guardar_datos(pedido):
    wb = load_workbook(filename='Ventas.xlsx')
    wb.active.append(pedido)
    print('\n','Abre bien el archivo','\n')
    wb.save('Ventas.xlsx')
    print("Carga exitosa de la venta!!")


def lista_productos():
    pedido.delete(0, tk.END)
    c = len(empanadas_de_carne)
    pedido.insert('0',c)


def cambiar_tarjeta_valor():
    if tarjeta_valor.get() == int(1):
        tarjeta_valor.set(0)
        print('Borrado boton tarjeta_valor')
    else:
        print('Boton no tildado..')
        print(tarjeta_valor.get())
        print(type(tarjeta_valor.get()))
        pass


def confirmar():
    m = messagebox.askokcancel(title='Confirmación', message='Desea confirmar el pedido?')
    if m:
        for caja in cajas:
            caja.delete(0, tk.END)
        facturacion.delete(0,tk.END)
        paga_con.delete(0,tk.END)
        vuelto.delete(0,tk.END)
        descuento.delete(0,tk.END)
    else:
        pass


def cancelar():
    m = messagebox.askokcancel(title='Cancelación', message='Desea cancelar el pedido?')
    if m:
        for caja in cajas:
            caja.delete(0, tk.END)
        facturacion.delete(0,tk.END)
        paga_con.delete(0,tk.END)
        vuelto.delete(0,tk.END)
        descuento.delete(0,tk.END)
    else:
        pass


def contenido(texto):
    try:
        gusto = float(texto.get())
    except:
        if texto.get() == '':
            gusto = float(0)
        else:
            messagebox.showinfo(title='Error', message='Ingrese un número válido.')
    return gusto


def suma():
    hora = time.asctime()
    if facturacion.get() == '':
        # Empanadas
        carne = contenido(texto_carne)
        pollo = contenido(texto_pollo)
        jq = contenido(texto_jq)
        ver = contenido(texto_verdura)
        cq = contenido(texto_cq)
        # Tartas
        tarta_jq = contenido(texto_tar_jq)
        tarta_puerro = contenido(texto_tar_puerro)
        tarta_beren = contenido(texto_tar_beren)
        tarta_acelga_queso = contenido(texto_tar_acelga_queso)
        tarta_acelga_cala = contenido(texto_tar_acelga_cala)
        tarta_cala = contenido(texto_tar_cala)
        tarta_zapa = contenido(texto_tar_zapa)
        tarta_jqch = contenido(texto_tar_jqch)
        # Bebidas
        gaseosa = contenido(texto_gaseosa)
        agua = contenido(texto_agua)
        cerveza = contenido(texto_cerveza)
        # Platos y comidas
        menu = contenido(texto_menu)
        menu_sin = contenido(texto_menu_sin)
        tortilla = contenido(texto_tortilla)
        ensalada = contenido(texto_ensalada)
        ensa_chica = contenido(texto_ensalada_chica)
        papas = contenido(texto_papas)
        omelette = contenido(texto_omelette)
        # Cafeteria
        cafe = contenido(texto_cafe)
        jarrito = contenido(texto_jarrito)
        cafe_leche = contenido(texto_cafe_leche)
        lagrima = contenido(texto_lagrima)
        te = contenido(texto_te)
        cafe_llevar = contenido(texto_cafe_llevar)
        alfa = contenido(texto_alfa)
        medialuna = contenido(texto_media)
        ensa_fruta = contenido(texto_ensa_fru)
        flan = contenido(texto_flan)

    else:
        facturacion.delete(0,tk.END)
        # Empanadas
        carne = contenido(texto_carne)
        pollo = contenido(texto_pollo)
        jq = contenido(texto_jq)
        ver = contenido(texto_verdura)
        cq = contenido(texto_cq)
        # Tartas
        tarta_jq = contenido(texto_tar_jq)
        tarta_puerro = contenido(texto_tar_puerro)
        tarta_beren = contenido(texto_tar_beren)
        tarta_acelga_queso = contenido(texto_tar_acelga_queso)
        tarta_acelga_cala = contenido(texto_tar_acelga_cala)
        tarta_cala = contenido(texto_tar_cala)
        tarta_zapa = contenido(texto_tar_zapa)
        tarta_jqch = contenido(texto_tar_jqch)
        # Bebidas
        gaseosa = contenido(texto_gaseosa)
        agua = contenido(texto_agua)
        cerveza = contenido(texto_cerveza)
        # Platos y comidas
        menu = contenido(texto_menu)
        menu_sin = contenido(texto_menu_sin)
        tortilla = contenido(texto_tortilla)
        ensalada = contenido(texto_ensalada)
        ensa_chica = contenido(texto_ensalada_chica)
        papas = contenido(texto_papas)
        omelette = contenido(texto_omelette)
        # Cafeteria
        cafe = contenido(texto_cafe)
        jarrito = contenido(texto_jarrito)
        cafe_leche = contenido(texto_cafe_leche)
        lagrima = contenido(texto_lagrima)
        te = contenido(texto_te)
        cafe_llevar = contenido(texto_cafe_llevar)
        alfa = contenido(texto_alfa)
        medialuna = contenido(texto_media)
        ensa_fruta = contenido(texto_ensa_fru)
        flan = contenido(texto_flan)

    descuento_clientes = contenido(descuento)
    total_empa = (carne+pollo+jq+ver+cq) * precios['empanadas']
    total_tarta = (tarta_jq +tarta_puerro+tarta_beren+tarta_acelga_cala+tarta_acelga_queso+tarta_cala+tarta_zapa+tarta_jqch) * precios['tartas']
    total_bebidas = gaseosa * precios['bebida']['gaseosa'] + agua * precios['bebida']['agua'] + cerveza * precios['bebida']['cerveza']
    total_cafeteria = (cafe * precios['cafeteria']['cafe_chico'] + jarrito * precios['cafeteria']['jarrito'] + cafe_leche * precios['cafeteria']['cafe_leche'] +
                      lagrima * precios['cafeteria']['lagrima'] + te * precios['cafeteria']['te'] + cafe_llevar * precios['cafeteria']['cafe_llevar'] +
                      alfa * precios['cafeteria']['alfa'] + medialuna * precios['cafeteria']['medialuna'] + ensa_fruta * precios['postre']['ensa_fruta'] +
                      flan * precios['postre']['flan'])
    total_comidas = (menu * precios['platos']['plato'] + menu_sin * precios['platos']['plato_sin_guar'] + tortilla * precios['platos']['tortilla'] +
                    ensalada * precios['platos']['ensalada'] + ensa_chica * precios['platos']['ensa_chica'] + papas * precios['platos']['porcion_papas']+
                    omelette * precios['platos']['omelette'])
    total_productos = total_empa + total_tarta + total_bebidas + total_cafeteria + total_comidas - descuento_clientes
    facturacion.insert("0", total_productos)

    pago_tarjeta = checkbox_clicked()
    lista_venta.append(total_productos)
    # EMPANADAS
    empanadas_de_carne.append(carne)
    empanadas_de_jq.append(jq)
    empanadas_de_pollo.append(pollo)
    empanadas_de_verduda.append(ver)
    empanadas_de_cq.append(cq)
    # TARTAS
    lista_tarta_jq.append(tarta_jq)
    lista_tarta_puerro.append(tarta_puerro)
    lista_tarta_beren.append(tarta_beren)
    lista_tarta_acelga_cala.append(tarta_acelga_cala)
    lista_tarta_acelga_queso.append(tarta_acelga_queso)
    lista_tarta_cala.append(tarta_cala)
    lista_tarta_zapa.append(tarta_zapa)
    lista_tarta_jqch.append(tarta_jqch)
    # PLATOS
    lista_menu.append(menu)
    lista_menu_sin.append(menu_sin)
    lista_tortilla.append(tortilla)
    lista_ensalada.append(ensalada)
    lista_ensalada_chica.append(ensa_chica)
    lista_por_papas.append(papas)
    lista_ome.append(omelette)
    # CAFETERIA
    lista_cafe.append(cafe)
    lista_jarrito.append(jarrito)
    lista_cafe_leche.append(cafe_leche)
    lista_lagrima.append(lagrima)
    lista_te.append(te)
    lista_cafe_llevar.append(cafe_llevar)
    lista_alfajor.append(alfa)
    lista_medialuna.append(medialuna)
    lista_ensa_fruta.append(ensa_fruta)
    lista_flan.append(flan)
    lista_gaseosa.append(gaseosa)
    lista_agua.append(agua)
    lista_cerveza.append(cerveza)

    al_excel = [hora, carne, jq, pollo, ver, cq, tarta_jq, tarta_puerro, tarta_beren, tarta_acelga_queso, tarta_acelga_cala, tarta_cala, tarta_zapa,
                tarta_jqch, gaseosa, agua, cerveza, menu, menu_sin, tortilla, ensalada, ensa_chica, papas, omelette, cafe, jarrito, cafe_leche,
                lagrima, te, cafe_llevar, alfa, medialuna, ensa_fruta, flan, descuento_clientes, pago_tarjeta, total_productos]
    print(al_excel)
    guardar_datos(al_excel)
    lista_productos()
    total_dia.delete(0,tk.END)


def venta_acumulada():
    venta = sum(lista_venta)
    total_dia.insert('0',venta)


def funcion_conjunta():
    confirmar()
    #guardar_datos(lista_excel)
    venta_acumulada()


def fun_vuelto():
    if vuelto.get() == '':
        if (float(facturacion.get()) > 0) & (float(paga_con.get()) >0):
            v = float(paga_con.get()) - float(facturacion.get())
            vuelto.insert('0', v)
        else:
            if float(paga_con.get() == '') | float(paga_con.get() == '0'):
                vuelto.insert('0', 0)
            else:
                messagebox.showinfo(title='Error', message='Ingrese montos de vuelto o facturación válidos.')
    else:
        vuelto.delete(0,tk.END)
        if (float(facturacion.get()) > 0) & (float(paga_con.get()) >0):
            v = float(paga_con.get()) - float(facturacion.get())
            vuelto.insert('0', v)
        else:
            if float(paga_con.get() == '') | float(paga_con.get() == '0'):
                vuelto.insert('0', 0)
            else:
                messagebox.showinfo(title='Error', message='Ingrese montos de vuelto o facturación válidos.')


def borrar_datos():
    for caja in cajas:
        caja.delete(0, tk.END)
    cambiar_tarjeta_valor()
    facturacion.delete(0,tk.END)
    paga_con.delete(0,tk.END)
    vuelto.delete(0,tk.END)
    descuento.delete(0,tk.END)


def checkbox_clicked():
    rta = tarjeta_valor.get()
    return rta


### EXCEL INICIAL
comprobar_archivo()

### APP DE ESCRITORIO
ventana = themes.ThemedTk()
ventana.set_theme('winxpblue') # Other 'plastik'
ventana.config(height=750, width=800)
ventana.title("Aplicación de ventas - Viejo Caballito Bar")
#ventana.iconbitmap(default='./images.ico')
### CHECKBOX
tarjeta_valor = tk.IntVar()
tarjeta = ttk.Checkbutton(text='Pago con tarjeta?', variable=tarjeta_valor, command=checkbox_clicked)
tarjeta.place(x=250, y=560)

## CAJAS
# CAJAS EMPANADAS Y TARTAS
pedido = ttk.Entry()
pedido.place(x=150, y=10,width=40)
pedido.insert(tk.END, '1')
texto_carne = ttk.Entry()
texto_carne.place(x=250, y=75)
texto_carne.insert(tk.END,"")
texto_pollo = ttk.Entry()
texto_pollo.place(x=250, y=100)
texto_pollo.insert(tk.END,"")
texto_jq = ttk.Entry()
texto_jq.place(x=250, y=125)
texto_jq.insert(tk.END,"")
texto_verdura = ttk.Entry()
texto_verdura.place(x=250, y=150)
texto_verdura.insert(tk.END,"")
texto_cq = ttk.Entry()
texto_cq.place(x=250, y=175)
texto_cq.insert(tk.END,"")
texto_tar_jq = ttk.Entry()
texto_tar_jq.place(x=250, y=225)
texto_tar_jq.insert(tk.END,"")
texto_tar_puerro = ttk.Entry()
texto_tar_puerro.place(x=250, y=250)
texto_tar_puerro.insert(tk.END,"")
texto_tar_beren = ttk.Entry()
texto_tar_beren.place(x=250, y=275)
texto_tar_beren.insert(tk.END,"")
texto_tar_acelga_queso = ttk.Entry()
texto_tar_acelga_queso.place(x=250, y=300)
texto_tar_acelga_queso.insert(tk.END,"")
texto_tar_acelga_cala = ttk.Entry()
texto_tar_acelga_cala.place(x=250, y=325)
texto_tar_acelga_cala.insert(tk.END,"")
texto_tar_cala = ttk.Entry()
texto_tar_cala.place(x=250, y=350)
texto_tar_cala.insert(tk.END,"")
texto_tar_zapa = ttk.Entry()
texto_tar_zapa.place(x=250, y=375)
texto_tar_zapa.insert(tk.END,"")
texto_tar_jqch = ttk.Entry()
texto_tar_jqch.place(x=250, y=400)
texto_tar_jqch.insert(tk.END,"")
# CAJAS BEBIDAS
texto_gaseosa = ttk.Entry()
texto_gaseosa.place(x=250, y=450)
texto_gaseosa.insert(tk.END,"")
texto_agua = ttk.Entry()
texto_agua.place(x=250, y=475)
texto_agua.insert(tk.END,"")
texto_cerveza = ttk.Entry()
texto_cerveza.place(x=250, y=500)
texto_cerveza.insert(tk.END,"")
# CAJAS PLATOS Y COMIDAS
texto_menu = ttk.Entry()
texto_menu.place(x=600, y=75)
texto_menu.insert(tk.END,"")
texto_menu_sin = ttk.Entry()
texto_menu_sin.place(x=600, y=100)
texto_menu_sin.insert(tk.END,"")
texto_tortilla = ttk.Entry()
texto_tortilla.place(x=600, y=125)
texto_tortilla.insert(tk.END,"")
texto_ensalada = ttk.Entry()
texto_ensalada.place(x=600, y=150)
texto_ensalada.insert(tk.END,"")
texto_ensalada_chica = ttk.Entry()
texto_ensalada_chica.place(x=600, y=175)
texto_ensalada_chica.insert(tk.END,"")
texto_papas = ttk.Entry()
texto_papas.place(x=600, y=200)
texto_papas.insert(tk.END,"")
texto_omelette = ttk.Entry()
texto_omelette.place(x=600, y=225)
texto_omelette.insert(tk.END,"")
# CAJAS CAFETERIA
texto_cafe = ttk.Entry()
texto_cafe.place(x=600, y=275)
texto_cafe.insert(tk.END,"")
texto_jarrito = ttk.Entry()
texto_jarrito.place(x=600, y=300)
texto_jarrito.insert(tk.END,"")
texto_cafe_leche = ttk.Entry()
texto_cafe_leche.place(x=600, y=325)
texto_cafe_leche.insert(tk.END,"")
texto_lagrima = ttk.Entry()
texto_lagrima.place(x=600, y=350)
texto_lagrima.insert(tk.END,"")
texto_te = ttk.Entry()
texto_te.place(x=600, y=375)
texto_te.insert(tk.END,"")
texto_cafe_llevar = ttk.Entry()
texto_cafe_llevar.place(x=600, y=400)
texto_cafe_llevar.insert(tk.END,"")
texto_alfa = ttk.Entry()
texto_alfa.place(x=600, y=425)
texto_alfa.insert(tk.END,"")
texto_media = ttk.Entry()
texto_media.place(x=600, y=450)
texto_media.insert(tk.END,"")
texto_ensa_fru = ttk.Entry()
texto_ensa_fru.place(x=600, y=475)
texto_ensa_fru.insert(tk.END,"")
texto_flan = ttk.Entry()
texto_flan.place(x=600, y=500)
texto_flan.insert(tk.END,"")

# CAJAS FACTURACION/VUELTO/PAGO
facturacion = ttk.Entry()
facturacion.place(x=250, y=620)
facturacion.insert(tk.END,'')
paga_con = ttk.Entry()
paga_con.place(x=250, y=650)
paga_con.insert(tk.END, '')
vuelto = ttk.Entry()
vuelto.place(x=250, y=680)
vuelto.insert(tk.END, '')
descuento = ttk.Entry()
descuento.place(x=250, y=590)
descuento.insert(tk.END,'')
total_dia = ttk.Entry()
total_dia.place(x=400, y=10,width=90)
total_dia.insert(tk.END,'')

cajas = [texto_carne, texto_pollo, texto_jq, texto_verdura, texto_cq, texto_tar_jq, texto_tar_puerro, texto_tar_beren, texto_tar_acelga_queso,
         texto_tar_acelga_cala, texto_tar_cala, texto_tar_zapa, texto_tar_jqch, texto_gaseosa, texto_agua, texto_cerveza, texto_menu, texto_menu_sin,
         texto_tortilla, texto_ensalada, texto_ensalada_chica, texto_papas, texto_omelette, texto_cafe, texto_jarrito, texto_cafe_leche, texto_lagrima,
         texto_te,texto_cafe_llevar, texto_alfa, texto_media, texto_ensa_fru, texto_flan]

## ETIQUETAS
# ETIQUETAS EMPANADAS Y TARTAS
ttk.Label(text=f'Número de pedido:').place(x=28, y=10)
ttk.Label(text="EMPANADAS").place(x=25, y=50)
lab_carne=ttk.Label(text='CARNE: ')
lab_carne.place(x=25, y=75)
lab_pollo = ttk.Label(text='POLLO: ')
lab_pollo.place(x=25, y=100)
lab_jq = ttk.Label(text='JAMON Y QUESO: ')
lab_jq.place(x=25, y=125)
lab_ver = ttk.Label(text='VERDURA: ')
lab_ver.place(x=25, y=150)
lab_cq = ttk.Label(text='CEBOLLA Y QUESO: ')
lab_cq.place(x=25, y=175)
ttk.Label(text="TARTAS").place(x=25, y=200)
lab_tar_jq = ttk.Label(text='JAMON Y QUESO: ')
lab_tar_jq.place(x=25, y=225)
lab_tar_puerro = ttk.Label(text='PUERRO Y QUESO: ')
lab_tar_puerro.place(x=25, y=250)
lab_tar_beren = ttk.Label(text='BERENJENA Y QUESO: ')
lab_tar_beren.place(x=25, y=275)
lab_tar_acelga_queso = ttk.Label(text='ACELGA Y QUESO: ')
lab_tar_acelga_queso.place(x=25, y=300)
lab_tar_acelga_cala = ttk.Label(text='ACELGA Y CALABAZA: ')
lab_tar_acelga_cala.place(x=25, y=325)
lab_tar_calabaza = ttk.Label(text='CALABAZA Y QUESO: ')
lab_tar_calabaza.place(x=25, y=350)
lab_tar_zapa = ttk.Label(text='ZAPALLITO: ')
lab_tar_zapa.place(x=25, y=375)
lab_tar_jqch = ttk.Label(text='JAMON Y QUESO, CEBOLLA: ')
lab_tar_jqch.place(x=25, y=400)
# BEBIDAS
ttk.Label(text="BEBIDAS").place(x=25, y=425)
lab_gaseosa=ttk.Label(text='GASEOSA: ')
lab_gaseosa.place(x=25, y=450)
lab_agua=ttk.Label(text='AGUA: ')
lab_agua.place(x=25, y=475)
lab_cerveza=ttk.Label(text='CERVEZA: ')
lab_cerveza.place(x=25, y=500)

# ETIQUETAS PLATOS
ttk.Label(text="PLATOS Y COMIDAS").place(x=400, y=50)
lab_menu=ttk.Label(text= 'PLATO DEL DIA: ')
lab_menu.place(x=400, y=75)
lab_menu_sin = ttk.Label(text='PLATO DEL DIA S/GUARNICION: ')
lab_menu_sin.place(x=400, y=100)
lab_tortilla=ttk.Label(text='TORTILLAS: ')
lab_tortilla.place(x=400, y=125)
lab_ensalada=ttk.Label(text='ENSALADA: ')
lab_ensalada.place(x=400, y=150)
lab_ensalada=ttk.Label(text='ENSALADA CHICA: ')
lab_ensalada.place(x=400, y=175)
lab_papas=ttk.Label(text='PORCIONES DE PAPAS: ')
lab_papas.place(x=400, y=200)
lab_ome=ttk.Label(text='OMELETTE: ')
lab_ome.place(x=400, y=225)
ttk.Label(text="CAFETERIA Y BEBIDAS").place(x=400, y=250)
lab_cafe=ttk.Label(text='CAFE CHICO: ')
lab_cafe.place(x=400, y=275)
lab_cafe=ttk.Label(text='JARRITO: ')
lab_cafe.place(x=400, y=300)
lab_cafe=ttk.Label(text='CAFE CON LECHE: ')
lab_cafe.place(x=400, y=325)
lab_cafe=ttk.Label(text='LAGRIMA: ')
lab_cafe.place(x=400, y=350)
lab_cafe=ttk.Label(text='TE: ')
lab_cafe.place(x=400, y=375)
lab_cafe=ttk.Label(text='CAFE PARA LLEVAR: ')
lab_cafe.place(x=400, y=400)
# ETIQUETA OTROS
lab_alfajores=ttk.Label(text='ALFAJORES: ')
lab_alfajores.place(x=400, y=425)
lab_medialunas=ttk.Label(text='MEDIALUNAS: ')
lab_medialunas.place(x=400, y=450)
lab_ensalada_fruta=ttk.Label(text='ENSALADAS DE FRUTA: ')
lab_ensalada_fruta.place(x=400, y=475)
lab_ensalada_fruta=ttk.Label(text='FLAN: ')
lab_ensalada_fruta.place(x=400, y=500)

# ETIQUETAS FACTURACION/VUELTO/PAGO
lab_fact = ttk.Label(text='Total a pagar: ')
lab_fact.place(x=25, y=620)
lab_paga = ttk.Label(text='Cliente paga con: ')
lab_paga.place(x=25, y=650)
lab_vuelto = ttk.Label(text='Vuelto a dar: ')
lab_vuelto.place(x=25, y=680)
lab_desc = ttk.Label(text='Descuento a clientes: ')
lab_desc.place(x=25, y=590)
lab_total = ttk.Label(text='VENTA ACUMULADA: ')
lab_total.place(x=270, y=10)


## BOTONES - MESSAGE BOX
ingresar = ttk.Button(text='Aceptar', command=funcion_conjunta)
cancelar = ttk.Button(text='Cancelar', command=cancelar)
ingresar.place(x=550, y=680)
cancelar.place(x=625, y=680)
boton_calcular = ttk.Button(text='Calcular venta', command=suma)
boton_calcular.place(x=400,y=620)
boton_vuelto = ttk.Button(text='Calcular vuelto', command=fun_vuelto)
boton_vuelto.place(x=400, y=650)
boton_borrar = ttk.Button(text='Borrar datos', command=borrar_datos)
boton_borrar.place(x=700, y=680)

ventana.mainloop()

