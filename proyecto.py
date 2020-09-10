"""
App de escritorio - Pedidos
"""
import tkinter as tk
from tkinter import messagebox, ttk
import os, time
import xlwings as xw
from openpyxl import Workbook, load_workbook

empanadas_de_carne = []
empanadas_de_pollo = []
empanadas_de_jq = []
empanadas_de_verduda = []


def comprobar_archivo():
    existe = os.path.exists('Ventas.xlsx')
    if existe:
        wb = load_workbook(filename='Ventas.xlsx')
        ws = wb.active
        print('Apertura del exitosa del archivo.')
    else:
        wb = Workbook()
        ws = wb.active
        titulo = ('Hora',"Carne", 'JQ', 'Pollo', 'Verdura', 'Total')
        ws.append(titulo)
        wb.save(filename='Ventas.xlsx')
        print('Creacion exitosa del archivo')


def guardar_datos(pedido):
    wb = load_workbook(filename='Ventas.xlsx')
    wb.active.append(pedido)
    wb.save('Ventas.xlsx')    
    print("Carga exitosa de la venta!!")


def lista_productos():
    pedido.delete(0, tk.END)
    c = len(empanadas_de_carne)
    pedido.insert('0',c)
    

def confirmar():
    m = messagebox.askokcancel(title='Confirmación', message='Desea confirmar el pedido?')
    if m:
        texto_carne.delete(0,tk.END)
        texto_pollo.delete(0,tk.END)
        texto_jq.delete(0,tk.END)
        texto_verdura.delete(0,tk.END)
        facturacion.delete(0,tk.END)
        paga_con.delete(0,tk.END)
        vuelto.delete(0,tk.END)
    else:
        pass


def cancelar():
    m = messagebox.askokcancel(title='Cancelación', message='Desea cancelar el pedido?')
    if m:
        texto_carne.delete(0,tk.END)
        texto_pollo.delete(0,tk.END)
        texto_jq.delete(0,tk.END)
        texto_verdura.delete(0,tk.END)
        facturacion.delete(0,tk.END)
        paga_con.delete(0,tk.END)
        vuelto.delete(0,tk.END)
    else:
        pass


def suma():
    hora = time.asctime()        
    if facturacion.get() == '':
        precio_empanada = float(60)
        try:
            carne = int(texto_carne.get())
        except:
            if texto_carne.get() == '':
                carne = int(0)
            else: 
                messagebox.showinfo(title='Error', message='Ingrese un numero valido.')
        try:
            pollo = int(texto_pollo.get())
        except:
            if texto_pollo.get() == '':
                pollo = int(0)
            else: 
                messagebox.showinfo(title='Error', message='Ingrese un numero valido.')
        try:
            jq = int(texto_jq.get())
        except:
            if texto_jq.get() == '':
                jq = int(0)
            else: 
                messagebox.showinfo(title='Error', message='Ingrese un numero valido.')
        try:
            ver = int(texto_verdura.get())
        except:
            if texto_verdura.get() == '':
                ver = int(0)
            else: 
                messagebox.showinfo(title='Error', message='Ingrese un numero valido.')
        
        total_empa = (carne+pollo+jq+ver) * precio_empanada
        facturacion.insert("0", total_empa)
    else:
        facturacion.delete(0,tk.END)
        precio_empanada = float(60)
        try:
            carne = int(texto_carne.get())
        except:
            if texto_carne.get() == '':
                carne = int(0)
            else: 
                messagebox.showinfo(title='Error', message='Ingrese un numero valido.')
        try:
            pollo = int(texto_pollo.get())
        except:
            if texto_pollo.get() == '':
                pollo = int(0)
            else: 
                messagebox.showinfo(title='Error', message='Ingrese un numero valido.')
        try:
            jq = int(texto_jq.get())
        except:
            if texto_jq.get() == '':
                jq = int(0)
            else: 
                messagebox.showinfo(title='Error', message='Ingrese un numero valido.')
        try:
            ver = int(texto_verdura.get())
        except:
            if texto_verdura.get() == '':
                ver = int(0)
            else: 
                messagebox.showinfo(title='Error', message='Ingrese un numero valido.')
        
        total_empa = (carne+pollo+jq+ver) * precio_empanada
        facturacion.insert("0", total_empa)

    empanadas_de_carne.append(carne)
    empanadas_de_jq.append(jq)
    empanadas_de_pollo.append(pollo)
    empanadas_de_verduda.append(ver)
    al_excel = [hora, carne, jq, pollo, ver, total_empa]
    guardar_datos(al_excel)
    lista_productos()

def fun_vuelto():    
    if vuelto.get() == '':
        if (float(facturacion.get()) > 0) & (float(paga_con.get()) >0):
            v = float(paga_con.get()) - float(facturacion.get())
            vuelto.insert('0', v)
        else:
            if float(paga_con.get() == '') | float(paga_con.get() == '0'):
                vuelto.insert('0', 0)
            else:
                messagebox.showinfo(title='Error', message='Ingrese montos de vuelto o facturacion validos.')
    else:
        vuelto.delete(0,tk.END)
        if (float(facturacion.get()) > 0) & (float(paga_con.get()) >0):
            v = float(paga_con.get()) - float(facturacion.get())
            vuelto.insert('0', v)
        else:
            if float(paga_con.get() == '') | float(paga_con.get() == '0'):
                vuelto.insert('0', 0)
            else:
                messagebox.showinfo(title='Error', message='Ingrese montos de vuelto o facturacion validos.')

### EXCEL INICIAL
comprobar_archivo()

### APP DE ESCRITORIO
ventana = tk.Tk()
ventana.config(height=400, width=550)
ventana.title("Aplicación de ventas - Viejo Caballito")

# CAJAS
pedido = ttk.Entry()
pedido.place(x=150, y=5,width=40)
pedido.insert(tk.END, '')
texto_carne = ttk.Entry()
texto_carne.place(x=300, y=50)
texto_carne.insert(tk.END,"0")
texto_pollo = ttk.Entry()
texto_pollo.place(x=300, y=75)
texto_pollo.insert(tk.END,"0")
texto_jq = ttk.Entry()
texto_jq.place(x=300, y=100)
texto_jq.insert(tk.END,"0")
texto_verdura = ttk.Entry()
texto_verdura.place(x=300, y=125)
texto_verdura.insert(tk.END,"0")

facturacion = ttk.Entry()
facturacion.place(x=300, y=240)
facturacion.insert(tk.END,'')

paga_con = ttk.Entry()
paga_con.place(x=300, y=270)
paga_con.insert(tk.END, '')
vuelto = ttk.Entry()
vuelto.place(x=300, y=300)
vuelto.insert(tk.END, '')

# BOTONES - MESSAGE BOX
ingresar = ttk.Button(text='Aceptar', command=confirmar)
cancelar = ttk.Button(text='Cancelar', command=cancelar)
ingresar.place(x=25, y=350)
cancelar.place(x=300, y=350)

boton_calcular = ttk.Button(text='Calcular venta', command=suma)
boton_calcular.place(x=450,y=240)

boton_vuelto = ttk.Button(text='Calcular vuelto', command=fun_vuelto)
boton_vuelto.place(x=450, y=300)

# ETIQUETAS
ttk.Label(text=f'Numero de pedido:').place(x=28, y=5)
ttk.Label(text=" Ingrese un nuevo pedido: ").place(x=25, y=26)
lab_carne=ttk.Label(text='Cantidad de empanadas de carne: ')
lab_carne.place(x=25, y=50)
lab_pollo = ttk.Label(text='Cantidad de empanadas de pollo: ')
lab_pollo.place(x=25, y=75)
lab_jq = ttk.Label(text='Cantidad de empanadas de jamon y queso: ')
lab_jq.place(x=25, y=100)
lab_ver = ttk.Label(text='Cantidad de empanadas de verduda: ')
lab_ver.place(x=25, y=125)

lab_fact = ttk.Label(text='Total a pagar: ')
lab_fact.place(x=25, y=240)
lab_paga = ttk.Label(text='Cliente paga con: ')
lab_paga.place(x=25, y=270)
lab_vuelto = ttk.Label(text='Vuelto a dar: ')
lab_vuelto.place(x=25, y=300)


ventana.mainloop()

