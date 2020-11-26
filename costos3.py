from tkinter import *
from tkinter import ttk
import tkinter as tk
import os, time
from openpyxl import Workbook, load_workbook


def comprobar_archivo():
    existe = os.path.exists('Costos.xlsx')
    if existe:
        wb = load_workbook(filename='Costos.xlsx')
        ws_costos = wb['Dato Costos']
        ws_obs = wb['Observaciones']
        ws_uni = wb['Unidades']
        print('Apertura exitosa del archivo.')
    else:
        wb = Workbook()
        ws_costos = wb.create_sheet('Dato Costos',0)
        ws_obs = wb.create_sheet('Observaciones',1)
        ws_uni = wb.create_sheet('Unidades',2)
        titulo_costos = ('Hora transacción','Pollo','Carne P.','Filet', 'Beef', 'Bondiola', 'Jamon', 'Nalga', 'Manteca','Aceite', 'Harina', 'Fideos', 'Arroz', 'Pure Tom.', 'Caldos', 'Pan', 'Medialuna', 'Soda G', 'Soda C', 'Cerveza', 'Agua C', 'Agua G', 'Gaseosa', 'Pan Rallado', 'Cebolla','Acelga', 'Papa', 'Calabaza', 'Zapallito', 'Berenjena', 'Puerro', 'Morron', 'Tomate', 'Verdeo', 'Zanahoria', 'Ajo', 'Pechuga', 'Batata','Limon','Pimenton','Pimienta','Provenzal','Nuez','Laurel','Tapas','Noquis', 'Huevos','Te','Azucar','Yerba','Edulcorante','Cafe Capsula','Cafe Molido','Cafe Kilo','Vasos Cafe','Alquiler','Luz','AYSA','Telefono','ABL','Diario','Fumig.','Deterg.','Monotr.','Gas','Tarj','IIBB','Otros','Flox','Carton','Sulfito','Film','Servilletas Mesa','Servilletas Cocina','Cubiertos','Bolsas Residuos','Bolsas Pedidos', 'Remos','Papel Higienico','Platos','Band. Ensa','Band. Pure','Band. tapa','Papel Alum.','Detergente','Lavandina','Perfume','Esponja','Jabon','Desengrasante','Alcohol','Otros','Tarjeta?','Total Fijos y Variables')
        ws_costos.append(titulo_costos)
        titulos_obs = ('Hora transacción', 'Alquiler','Luz','AYSA','Telefono','ABL','Diario','Fumig.','Deterg.','Monotr.','Gas','Tarjeta','IIBB','Otros','Tarjeta?', 'Total Costos Fijos')
        ws_obs.append(titulos_obs)
        titulos_uni =('Hora transacción', 'Cantidad Pollo', 'Costo Carne Picada', 'Cantidad Filet', 'Cantidad Beef', 'Cantidad Bondiola', 'Cantidad Jamon', 'Cantidad Nalga', 'Cantidad Manteca', 'Cantidad Aceite', 'Cantidad Harina', 'Cantidad Fideos', 'Cantidad Arroz', 'Cantidad pure tomate', 'Cantidad Caldos', 'Cantidad tomate', 'Costo tomate', 'Cantidad lechuga', 'Costo lechuga', 'Cantidad yogur', 'Costo yogur', 'Cantidad agua', 'Costo agua', 'Cantidad nalga', 'Costo nalga', 'Cantidad empleados', 'Costo empleados', 'Cantidad acelga', 'Costo acelga',  'Cantidad huevos', 'Pan', 'Medialuna', 'Soda G', 'Soda C', 'Cerveza', 'Agua C', 'Agua G', 'Gaseosa', 'Pan Rallado', 'Cebolla','Acelga', 'Papa', 'Calabaza', 'Zapallito', 'Berenjena', 'Puerro', 'Morron', 'Tomate', 'Verdeo', 'Zanahoria', 'Ajo', 'Pechuga', 'Batata','Limon','Pimenton','Pimienta','Provenzal','Nuez','Laurel','Tapas','Noquis', 'Huevos','Te','Azucar','Yerba','Edulcorante','Cafe Capsula','Cafe Molido','Cafe Kilo','Vasos Cafe')
        ws_uni.append(titulos_uni)
        wb.save(filename='Costos.xlsx')
        print('Creación exitosa del archivo')


def guardar_datos_costos(pedido):
    wb = load_workbook(filename='Costos.xlsx')
    wb['Dato Costos'].append(pedido)
    wb.save('Costos.xlsx')
    print("Carga exitosa del costo!!")


def guardar_datos_obs(info):
    wb = load_workbook(filename='Costos.xlsx')
    wb['Observaciones'].append(info)
    wb.save('Costos.xlsx')
    print("Carga exitosa de las observaciones!!")


def guardar_datos_unidades(info):
    wb = load_workbook(filename='Costos.xlsx')
    wb['Unidades'].append(info)
    wb.save('Costos.xlsx')
    print("Carga exitosa de las unidades!!")


def cambiar_tarjeta_valor():
    if tarjeta_valor.get() == int(1):
        tarjeta_valor.set(0)
        print('Borrado boton tarjeta_valor')
    else:
        print('Boton no tildado..')
        print(tarjeta_valor.get())
        print(type(tarjeta_valor.get()))
        pass


def contenido(caja_precio, caja_cantidad, total):
    try:
        cantidad = int(caja_cantidad.get())
        precio = int(caja_precio.get())
        variable = int(caja_precio.get()) * int(caja_cantidad.get())
        total.insert('0',variable)
    except:
        if (caja_precio.get() == '') | (caja_cantidad.get() == ''):
            variable = 0
            cantidad = 0
            precio = 0
        else:
            messagebox.showinfo(title='Error', message='Ingresar un número válido')
    return variable, cantidad, precio


def contenido_fijos(caja_precio, caja_obs):
    try:
        variable = int(caja_precio.get())
        observacion = caja_obs.get()
    except:
        if caja_precio.get() == '':
            variable = 0
            observacion = ''
        else:
            messagebox.showinfo(title='Error', message='Ingresar un número válido')
    return variable, observacion


def mult():
    hora = time.asctime()
    if caja_total.get() == '':
        ## COMPLETA VALORES CON FUNCION 'CONTENIDO'
        # COSTOS VARIABLES
        pollo, q_pollo, p_pollo = contenido(caja_p_pollo, caja_q_pollo, caja_t_pollo)
        picada, q_picada, p_picada = contenido(caja_p_picada, caja_q_picada, caja_t_picada)
        filet, q_filet, p_filet = contenido(caja_p_filet, caja_q_filet, caja_t_filet)
        beef, q_beef, p_beef = contenido(caja_p_beef, caja_q_beef, caja_t_beef)
        bondiola, q_bondiola, p_bondiola = contenido(caja_p_bondiola, caja_q_bondiola, caja_t_bondiola)
        jamon, q_jamon, p_jamon = contenido(caja_p_jamon, caja_q_jamon, caja_t_jamon)
        nalga, q_nalga, p_nalga = contenido(caja_p_nalga, caja_q_nalga, caja_t_nalga)
        manteca, q_manteca, p_manteca = contenido(caja_p_manteca, caja_q_manteca, caja_t_manteca)
        aceite, q_aceite, p_aceite = contenido(caja_p_aceite, caja_q_aceite, caja_t_aceite)
        harina, q_harina, p_harina = contenido(caja_p_harina, caja_q_harina, caja_t_harina)
        fideos, q_fideos, p_fideos = contenido(caja_p_fideos, caja_q_fideos, caja_t_fideos)
        arroz, q_arroz, p_arroz = contenido(caja_p_arroz, caja_q_arroz, caja_t_arroz)
        pure_tom, q_pure_tom, p_pure_tom = contenido(caja_p_pure_tom, caja_q_pure_tom, caja_t_pure_tom)
        caldos, q_caldos, p_caldos = contenido(caja_p_caldos, caja_q_caldos, caja_t_caldos)
        pan, q_pan, p_pan = contenido(caja_p_pan, caja_q_pan, caja_t_pan)
        medialuna, q_medialuna, p_medialuna = contenido(caja_p_medial, caja_q_medial, caja_t_medial)
        sodag, q_sodag, p_sodag = contenido(caja_p_sodag, caja_q_sodag, caja_t_sodag)
        sodac, q_sodac, p_sodac = contenido(caja_p_sodac, caja_q_sodac, caja_t_sodac)
        cerveza, q_cerveza, p_cerveza = contenido(caja_p_cerveza, caja_q_cerveza, caja_t_cerveza)
        aguac, q_aguac, p_aguac = contenido(caja_p_aguac, caja_q_aguac, caja_t_aguac)
        aguag, q_aguag, p_aguag = contenido(caja_p_aguag, caja_q_aguag, caja_t_aguag)
        gaseosa, q_gaseosa, p_gaseosa = contenido(caja_p_gaseosa, caja_q_gaseosa, caja_t_gaseosa)
        pan_rallado, q_pan_ra, p_pan_ra = contenido(caja_p_pan_rallado, caja_q_pan_rallado, caja_t_pan_rallado)
        cebolla, q_cebolla, p_cebolla = contenido(caja_p_cebolla, caja_q_cebolla, caja_t_cebolla)
        acelga, q_acelga, p_acelga = contenido(caja_p_acelga, caja_q_acelga, caja_t_acelga)
        papa, q_papa, p_papa = contenido(caja_p_papa, caja_q_papa, caja_t_papa)
        cala, q_cala, p_cala = contenido(caja_p_cala, caja_q_cala, caja_t_cala)
        zapallito, q_zapallito, p_zapallito = contenido(caja_p_zapa, caja_q_zapa, caja_t_zapa)
        beren, q_beren, p_beren = contenido(caja_p_beren, caja_q_beren, caja_t_beren)
        puerro, q_puerro, p_puerro = contenido(caja_p_puerro, caja_q_puerro, caja_t_puerro)
        morron, q_morron, p_morron = contenido(caja_p_morron, caja_q_morron, caja_t_morron)
        tomate, q_tomate, p_tomate = contenido(caja_p_tomate, caja_q_tomate, caja_t_tomate)
        verdeo, q_verdeo, p_verdeo = contenido(caja_p_verdeo, caja_q_verdeo, caja_t_verdeo)
        zanahoria, q_zana, p_zana = contenido(caja_p_zana, caja_q_zana, caja_t_zana)
        ajo, q_ajo, p_ajo = contenido(caja_p_ajo, caja_q_ajo, caja_t_ajo)
        pechuga, q_pechuga, p_pechuga = contenido(caja_p_pechuga, caja_q_pechuga, caja_t_pechuga)
        batata, q_batata, p_batata = contenido(caja_p_batata, caja_q_batata, caja_t_batata)
        limon, q_limon, p_limon = contenido(caja_p_limon, caja_q_limon, caja_t_limon)
        oregano, q_oregano, p_oregano = contenido(caja_p_oregano, caja_q_oregano, caja_t_oregano)
        pimenton, q_pimenton, p_pimenton = contenido(caja_p_pimenton, caja_q_pimenton, caja_t_pimenton)
        pimienta, q_pimienta, p_pimienta = contenido(caja_p_pimienta, caja_q_pimienta, caja_t_pimienta)
        provenzal, q_provenzal, p_provenzal = contenido(caja_p_provenzal, caja_q_provenzal, caja_t_provenzal)
        nuez, q_nuez, p_nuez = contenido(caja_p_nuez, caja_q_nuez, caja_t_nuez)
        laurel, q_laurel, p_laurel = contenido(caja_p_laurel, caja_q_laurel, caja_t_laurel)
        tapas, q_tapas, p_tapas = contenido(caja_p_tapas, caja_q_tapas, caja_t_tapas)
        noquiz, q_noquiz, p_noquiz = contenido(caja_p_noquiz, caja_q_noquiz, caja_t_noquiz)
        huevos, q_huevos, p_huevos = contenido(caja_p_huevos, caja_q_huevos, caja_t_huevos)
        te, q_te, p_te = contenido(caja_p_te, caja_q_te, caja_t_te)
        azucar, q_azucar, p_azucar = contenido(caja_p_azucar, caja_q_azucar, caja_t_azucar)
        yerba, q_yerba, p_yerba = contenido(caja_p_yerba, caja_q_yerba, caja_t_yerba)
        edulco, q_edulco, p_edulco = contenido(caja_p_edulco, caja_q_edulco, caja_t_edulco)
        capsula, q_capsula, p_capsula = contenido(caja_p_capsula, caja_q_capsula, caja_t_capsula)
        molido, q_molido, p_molido = contenido(caja_p_molido, caja_q_molido, caja_t_molido)
        cafe_kilo, q_cafe_kilo, p_cafe_kilo = contenido(caja_p_cafe_kilo, caja_q_cafe_kilo, caja_t_cafe_kilo)
        vasos_cafe, q_vasos_cafe, p_vasos_cafe = contenido(caja_p_vasos_cafe, caja_q_vasos_cafe, caja_t_vasos_cafe)


        # COSTOS FIJOS
        alquiler = contenido_fijos(caja_p_alquiler,caja_obs_alquiler)[0]
        luz = contenido_fijos(caja_p_luz, caja_obs_luz)[0]
        agua_servicio = contenido_fijos(caja_p_agua_servicio, caja_obs_agua)[0]
        telefono = contenido_fijos(caja_p_telefono, caja_obs_telefono)[0]
        abl = contenido_fijos(caja_p_abl, caja_obs_abl)[0]
        diario = contenido_fijos(caja_p_diario, caja_obs_diario)[0]
        fumigador = contenido_fijos(caja_p_fumigador, caja_obs_fumigador)[0]
        detergente = contenido_fijos(caja_p_detergente, caja_obs_fumigador)[0]
        monotributo = contenido_fijos(caja_p_monotributo, caja_obs_monotributo)[0]
        gas = contenido_fijos(caja_p_gas, caja_obs_gas)[0]
        tarjeta = contenido_fijos(caja_p_tarjeta, caja_obs_tarjeta)[0]
        iibb = contenido_fijos(caja_p_iibb, caja_obs_iibb)[0]
        otros = contenido_fijos(caja_p_otros, caja_obs_otros)[0]
        # COSTOS FIJOS 2
        flox, q_flox, p_flox = contenido(caja_p_flox, caja_q_flox, caja_t_flox)
        carton, q_carton, p_carton = contenido(caja_p_carton, caja_q_carton, caja_t_carton)
        sulfito, q_sulfito, p_sulfito = contenido(caja_p_sulfito, caja_q_sulfito, caja_t_sulfito)
        film, q_film, p_film = contenido(caja_p_film, caja_q_film, caja_t_film)
        serv_mesa, q_serv_mesa, p_serv_mesa = contenido(caja_p_serv_mesa, caja_q_serv_mesa, caja_t_serv_mesa)
        serv_cocina, q_serv_cocina, p_serv_cocina = contenido(caja_p_serv_cocina, caja_q_serv_cocina, caja_t_serv_cocina)
        cubiertos, q_cubiertos, p_cubiertos = contenido(caja_p_cubiertos, caja_q_cubiertos, caja_t_cubiertos)
        bolsas_res, q_bolsas_res, p_bolsas_res = contenido(caja_p_bolsas_residuos, caja_q_bolsas_residuos, caja_t_bolsas_residuos)
        bolsas_ped, q_bolsas_ped, p_bolsas_ped = contenido(caja_p_bolsas_pedidos, caja_q_bolsas_pedidos, caja_t_bolsas_pedidos)
        remos, q_remos, p_remos = contenido(caja_p_remos, caja_q_remos, caja_t_remos)
        higienico, q_higienico, p_higienico = contenido(caja_p_higienico, caja_q_higienico, caja_t_higienico)
        platos, q_platos, p_platos = contenido(caja_p_platos, caja_q_platos, caja_t_platos)
        band_ensa, q_band_ensa, p_band_ensa = contenido(caja_p_band_ensa, caja_q_band_ensa, caja_t_band_ensa)
        band_pure, q_band_pure, p_band_pure = contenido(caja_p_band_pure, caja_q_band_pure, caja_t_band_pure)
        band_tapa, q_band_tapa, p_band_tapa = contenido(caja_p_band_tapa, caja_q_band_tapa, caja_t_band_tapa)
        alum, q_alum, p_alum = contenido(caja_p_alum, caja_q_alum, caja_t_alum)
        detergente, q_detergente, p_detergente = contenido(caja_p_detergente, caja_q_detergente, caja_t_detergente)
        lavandina, q_lavandina, p_lavandina = contenido(caja_p_lavan, caja_q_lavan, caja_t_lavan)
        perfume, q_perfume, p_perfume = contenido(caja_p_perfume, caja_q_perfume, caja_t_perfume)
        esponja, q_esponja, p_esponja = contenido(caja_p_esponja, caja_q_esponja, caja_t_esponja)
        jabon, q_jabon, p_jabon = contenido(caja_p_jabon, caja_q_jabon, caja_t_jabon)
        desengrasante, q_desengrasante, p_desengrasante = contenido(caja_p_desengrasante, caja_q_desengrasante, caja_t_desengrasante)
        alcohol, q_alcohol, p_alcohol = contenido(caja_p_alcohol, caja_q_alcohol, caja_t_alcohol)


    else:
        # BORRA CASILLEROS COMPLETOS
        for caja in cajas:
            try:
                caja.delete(0, tk.END)
            except:
                pass

        cambiar_tarjeta_valor()

        # COMPLETA NUEVAMENTE VALORES
        # COSTOS VARIABLES
        pollo, q_pollo, p_pollo = contenido(caja_p_pollo, caja_q_pollo, caja_t_pollo)
        picada, q_picada, p_picada = contenido(caja_p_picada, caja_q_picada, caja_t_picada)
        filet, q_filet, p_filet = contenido(caja_p_filet, caja_q_filet, caja_t_filet)
        beef, q_beef, p_beef = contenido(caja_p_beef, caja_q_beef, caja_t_beef)
        bondiola, q_bondiola, p_bondiola = contenido(caja_p_bondiola, caja_q_bondiola, caja_t_bondiola)
        jamon, q_jamon, p_jamon = contenido(caja_p_jamon, caja_q_jamon, caja_t_jamon)
        nalga, q_nalga, p_nalga = contenido(caja_p_nalga, caja_q_nalga, caja_t_nalga)
        manteca, q_manteca, p_manteca = contenido(caja_p_manteca, caja_q_manteca, caja_t_manteca)
        aceite, q_aceite, p_aceite = contenido(caja_p_aceite, caja_q_aceite, caja_t_aceite)
        harina, q_harina, p_harina = contenido(caja_p_harina, caja_q_harina, caja_t_harina)
        fideos, q_fideos, p_fideos = contenido(caja_p_fideos, caja_q_fideos, caja_t_fideos)
        arroz, q_arroz, p_arroz = contenido(caja_p_arroz, caja_q_arroz, caja_t_arroz)
        pure_tom, q_pure_tom, p_pure_tom = contenido(caja_p_pure_tom, caja_q_pure_tom, caja_t_pure_tom)
        caldos, q_caldos, p_caldos = contenido(caja_p_caldos, caja_q_caldos, caja_t_caldos)
        pan, q_pan, p_pan = contenido(caja_p_pan, caja_q_pan, caja_t_pan)
        medialuna, q_medialuna, p_medialuna = contenido(caja_p_medial, caja_q_medial, caja_t_medial)
        sodag, q_sodag, p_sodag = contenido(caja_p_sodag, caja_q_sodag, caja_t_sodag)
        sodac, q_sodac, p_sodac = contenido(caja_p_sodac, caja_q_sodac, caja_t_sodac)
        cerveza, q_cerveza, p_cerveza = contenido(caja_p_cerveza, caja_q_cerveza, caja_t_cerveza)
        aguac, q_aguac, p_aguac = contenido(caja_p_aguac, caja_q_aguac, caja_t_aguac)
        aguag, q_aguag, p_aguag = contenido(caja_p_aguag, caja_q_aguag, caja_t_aguag)
        gaseosa, q_gaseosa, p_gaseosa = contenido(caja_p_gaseosa, caja_q_gaseosa, caja_t_gaseosa)
        pan_rallado, q_pan_ra, p_pan_ra = contenido(caja_p_pan_rallado, caja_q_pan_rallado, caja_t_pan_rallado)
        cebolla, q_cebolla, p_cebolla = contenido(caja_p_cebolla, caja_q_cebolla, caja_t_cebolla)
        acelga, q_acelga, p_acelga = contenido(caja_p_acelga, caja_q_acelga, caja_t_acelga)
        papa, q_papa, p_papa = contenido(caja_p_papa, caja_q_papa, caja_t_papa)
        cala, q_cala, p_cala = contenido(caja_p_cala, caja_q_cala, caja_t_cala)
        zapallito, q_zapallito, p_zapallito = contenido(caja_p_zapa, caja_q_zapa, caja_t_zapa)
        beren, q_beren, p_beren = contenido(caja_p_beren, caja_q_beren, caja_t_beren)
        puerro, q_puerro, p_puerro = contenido(caja_p_puerro, caja_q_puerro, caja_t_puerro)
        morron, q_morron, p_morron = contenido(caja_p_morron, caja_q_morron, caja_t_morron)
        tomate, q_tomate, p_tomate = contenido(caja_p_tomate, caja_q_tomate, caja_t_tomate)
        verdeo, q_verdeo, p_verdeo = contenido(caja_p_verdeo, caja_q_verdeo, caja_t_verdeo)
        zanahoria, q_zana, p_zana = contenido(caja_p_zana, caja_q_zana, caja_t_zana)
        ajo, q_ajo, p_ajo = contenido(caja_p_ajo, caja_q_ajo, caja_t_ajo)
        pechuga, q_pechuga, p_pechuga = contenido(caja_p_pechuga, caja_q_pechuga, caja_t_pechuga)
        batata, q_batata, p_batata = contenido(caja_p_batata, caja_q_batata, caja_t_batata)
        limon, q_limon, p_limon = contenido(caja_p_limon, caja_q_limon, caja_t_limon)
        oregano, q_oregano, p_oregano = contenido(caja_p_oregano, caja_q_oregano, caja_t_oregano)
        pimenton, q_pimenton, p_pimenton = contenido(caja_p_pimenton, caja_q_pimenton, caja_t_pimenton)
        pimienta, q_pimienta, p_pimienta = contenido(caja_p_pimienta, caja_q_pimienta, caja_t_pimienta)
        provenzal, q_provenzal, p_provenzal = contenido(caja_p_provenzal, caja_q_provenzal, caja_t_provenzal)
        nuez, q_nuez, p_nuez = contenido(caja_p_nuez, caja_q_nuez, caja_t_nuez)
        laurel, q_laurel, p_laurel = contenido(caja_p_laurel, caja_q_laurel, caja_t_laurel)
        tapas, q_tapas, p_tapas = contenido(caja_p_tapas, caja_q_tapas, caja_t_tapas)
        noquiz, q_noquiz, p_noquiz = contenido(caja_p_noquiz, caja_q_noquiz, caja_t_noquiz)
        huevos, q_huevos, p_huevos = contenido(caja_p_huevos, caja_q_huevos, caja_t_huevos)
        te, q_te, p_te = contenido(caja_p_te, caja_q_te, caja_t_te)
        azucar, q_azucar, p_azucar = contenido(caja_p_azucar, caja_q_azucar, caja_t_azucar)
        yerba, q_yerba, p_yerba = contenido(caja_p_yerba, caja_q_yerba, caja_t_yerba)
        edulco, q_edulco, p_edulco = contenido(caja_p_edulco, caja_q_edulco, caja_t_edulco)
        capsula, q_capsula, p_capsula = contenido(caja_p_capsula, caja_q_capsula, caja_t_capsula)
        molido, q_molido, p_molido = contenido(caja_p_molido, caja_q_molido, caja_t_molido)
        cafe_kilo, q_cafe_kilo, p_cafe_kilo = contenido(caja_p_cafe_kilo, caja_q_cafe_kilo, caja_t_cafe_kilo)
        vasos_cafe, q_vasos_cafe, p_vasos_cafe = contenido(caja_p_vasos_cafe, caja_q_vasos_cafe, caja_t_vasos_cafe)


        # COSTOS FIJOS
        alquiler = contenido_fijos(caja_p_alquiler,caja_obs_alquiler)[0]
        luz = contenido_fijos(caja_p_luz, caja_obs_luz)[0]
        agua_servicio = contenido_fijos(caja_p_agua_servicio, caja_obs_agua)[0]
        telefono = contenido_fijos(caja_p_telefono, caja_obs_telefono)[0]
        abl = contenido_fijos(caja_p_abl, caja_obs_abl)[0]
        diario = contenido_fijos(caja_p_diario, caja_obs_diario)[0]
        fumigador = contenido_fijos(caja_p_fumigador, caja_obs_fumigador)[0]
        detergente = contenido_fijos(caja_p_detergente, caja_obs_fumigador)[0]
        monotributo = contenido_fijos(caja_p_monotributo, caja_obs_monotributo)[0]
        gas = contenido_fijos(caja_p_gas, caja_obs_gas)[0]
        tarjeta = contenido_fijos(caja_p_tarjeta, caja_obs_tarjeta)[0]
        iibb = contenido_fijos(caja_p_iibb, caja_obs_iibb)[0]
        otros = contenido_fijos(caja_p_otros, caja_obs_otros)[0]
        # COSTOS FIJOS 2
        flox, q_flox, p_flox = contenido(caja_p_flox, caja_q_flox, caja_t_flox)
        carton, q_carton, p_carton = contenido(caja_p_carton, caja_q_carton, caja_t_carton)
        sulfito, q_sulfito, p_sulfito = contenido(caja_p_sulfito, caja_q_sulfito, caja_t_sulfito)
        film, q_film, p_film = contenido(caja_p_film, caja_q_film, caja_t_film)
        serv_mesa, q_serv_mesa, p_serv_mesa = contenido(caja_p_serv_mesa, caja_q_serv_mesa, caja_t_serv_mesa)
        serv_cocina, q_serv_cocina, p_serv_cocina = contenido(caja_p_serv_cocina, caja_q_serv_cocina, caja_t_serv_cocina)
        cubiertos, q_cubiertos, p_cubiertos = contenido(caja_p_cubiertos, caja_q_cubiertos, caja_t_cubiertos)
        bolsas_res, q_bolsas_res, p_bolsas_res = contenido(caja_p_bolsas_residuos, caja_q_bolsas_residuos, caja_t_bolsas_residuos)
        bolsas_ped, q_bolsas_ped, p_bolsas_ped = contenido(caja_p_bolsas_pedidos, caja_q_bolsas_pedidos, caja_t_bolsas_pedidos)
        remos, q_remos, p_remos = contenido(caja_p_remos, caja_q_remos, caja_t_remos)
        higienico, q_higienico, p_higienico = contenido(caja_p_higienico, caja_q_higienico, caja_t_higienico)
        platos, q_platos, p_platos = contenido(caja_p_platos, caja_q_platos, caja_t_platos)
        band_ensa, q_band_ensa, p_band_ensa = contenido(caja_p_band_ensa, caja_q_band_ensa, caja_t_band_ensa)
        band_pure, q_band_pure, p_band_pure = contenido(caja_p_band_pure, caja_q_band_pure, caja_t_band_pure)
        band_tapa, q_band_tapa, p_band_tapa = contenido(caja_p_band_tapa, caja_q_band_tapa, caja_t_band_tapa)
        alum, q_alum, p_alum = contenido(caja_p_alum, caja_q_alum, caja_t_alum)
        detergente, q_detergente, p_detergente = contenido(caja_p_detergente, caja_q_detergente, caja_t_detergente)
        lavandina, q_lavandina, p_lavandina = contenido(caja_p_lavan, caja_q_lavan, caja_t_lavan)
        perfume, q_perfume, p_perfume = contenido(caja_p_perfume, caja_q_perfume, caja_t_perfume)
        esponja, q_esponja, p_esponja = contenido(caja_p_esponja, caja_q_esponja, caja_t_esponja)
        jabon, q_jabon, p_jabon = contenido(caja_p_jabon, caja_q_jabon, caja_t_jabon)
        desengrasante, q_desengrasante, p_desengrasante = contenido(caja_p_desengrasante, caja_q_desengrasante, caja_t_desengrasante)
        alcohol, q_alcohol, p_alcohol = contenido(caja_p_alcohol, caja_q_alcohol, caja_t_alcohol)

    pago_tarjeta = checkbox_clicked()
    costos_varios = pollo + picada + filet + beef + bondiola + jamon + nalga + manteca + aceite + harina + fideos + arroz + pure_tom + caldos + pan + medialuna + sodag + sodac + cerveza + aguac + aguag + gaseosa + pan_rallado + cebolla + acelga + papa + cala + zapallito + beren + puerro + morron + tomate + verdeo + zanahoria + ajo + pechuga + batata + limon + oregano + pimenton + pimienta + provenzal + nuez + laurel + tapas + noquiz + huevos+ te + azucar + yerba + edulco + capsula + molido + cafe_kilo + vasos_cafe
    costos_fijos = alquiler + luz + agua_servicio + telefono + abl + diario + fumigador + detergente + monotributo + gas + tarjeta + iibb + otros + flox + carton + sulfito + film + serv_mesa + serv_cocina + cubiertos + bolsas_res + bolsas_ped + remos + higienico + platos + band_ensa + band_pure + band_tapa + alum + detergente + lavandina + perfume + esponja + jabon + desengrasante + alcohol
    facturacion = costos_varios + costos_fijos
    caja_total.insert('0',facturacion)

    # PASAJE A EXCEL
    al_excel = [hora, pollo, picada, filet, beef, bondiola, jamon, nalga, manteca, aceite, harina, fideos, arroz, pure_tom, caldos, pan, medialuna, sodag,
                sodac, cerveza, aguac, aguag, gaseosa, pan_rallado, cebolla, acelga, papa, cala, zapallito, beren, puerro, morron, tomate, verdeo,
                zanahoria, ajo, pechuga, batata, limon, oregano, pimenton, pimienta, provenzal, nuez, laurel, tapas, noquiz, huevos, te, azucar, yerba, edulco,
                capsula, molido, cafe_kilo, vasos_cafe, alquiler, luz, agua_servicio, telefono, abl, diario, fumigador, detergente, monotributo, gas, tarjeta,
                iibb, otros, flox, carton, sulfito, film, serv_mesa, serv_cocina, cubiertos, bolsas_res, bolsas_ped, remos, higienico, platos, band_ensa,
                band_pure, band_tapa, alum, detergente, lavandina, perfume, esponja, jabon, desengrasante, alcohol, pago_tarjeta, facturacion]
    al_excel_obs = [hora, contenido_fijos(caja_p_alquiler,caja_obs_alquiler)[1], contenido_fijos(caja_p_luz, caja_obs_luz)[1],
                    contenido_fijos(caja_p_agua_servicio, caja_obs_agua)[1], contenido_fijos(caja_p_telefono, caja_obs_telefono)[1],
                    contenido_fijos(caja_p_abl, caja_obs_abl)[1], contenido_fijos(caja_p_diario, caja_obs_diario)[1],
                    contenido_fijos(caja_p_fumigador, caja_obs_fumigador)[1], contenido_fijos(caja_p_detergente, caja_obs_detergente)[1],
                    contenido_fijos(caja_p_monotributo, caja_obs_monotributo)[1], contenido_fijos(caja_p_gas, caja_obs_gas)[1],
                    contenido_fijos(caja_p_tarjeta, caja_obs_tarjeta)[1], contenido_fijos(caja_p_iibb, caja_obs_iibb)[1],
                    contenido_fijos(caja_p_otros, caja_obs_otros)[1], pago_tarjeta,costos_fijos]
    al_excel_unidades = [hora, q_pollo, q_picada, q_filet, q_beef, q_bondiola, q_jamon, q_nalga, q_manteca, q_aceite, q_harina, q_fideos, q_arroz, q_pure_tom,
                         q_caldos, q_pan, q_medialuna, q_sodag, q_sodac, q_cerveza, q_aguac, q_aguag, q_gaseosa, q_pan, q_cebolla, q_acelga, q_papa, q_cala,
                         q_zapallito, q_beren, q_puerro, q_morron, q_tomate, q_verdeo, q_zana, q_ajo, q_pechuga, q_batata, q_limon, q_oregano, q_pimenton, q_pimienta,
                         q_provenzal, q_nuez, q_laurel, q_tapas, q_noquiz, q_huevos, q_te, q_azucar, q_yerba, q_edulco, q_capsula, q_molido, q_cafe_kilo,
                         q_vasos_cafe, q_flox, q_carton, q_sulfito, q_film, q_serv_mesa, q_serv_mesa, q_serv_cocina, q_cubiertos, q_bolsas_res, q_bolsas_ped,
                         q_remos, q_higienico, q_platos, q_band_ensa, q_band_pure, q_band_tapa, q_alum, q_detergente, q_lavandina, q_perfume, q_esponja,
                         q_jabon, q_desengrasante, q_alcohol]
    guardar_datos_costos(al_excel)
    guardar_datos_obs(al_excel_obs)
    guardar_datos_unidades(al_excel_unidades)


def confirmar():
    if caja_total == '':
        messagebox.showinfo(title='Error', message='Ingresar datos numéricos en el registro.')
    else:
        for caja in cajas:
            caja.delete(0, tk.END)

        caja_total.delete(0,tk.END)
        cambiar_tarjeta_valor()


def borrar_datos():
    for caja in cajas:
        caja.delete(0, tk.END)
    cambiar_tarjeta_valor()


def checkbox_clicked():
    rta = tarjeta_valor.get()
    return rta


### EXCEL INICIAL
comprobar_archivo()

### APP DE ESCRITORIO
root = Tk()
root.title('Viejo Caballito - Costos')
root.iconbitmap('./images.ico')
root.geometry('1400x800')

my_notebook = ttk.Notebook(root)
my_notebook.pack()

my_frame1 = Frame(my_notebook, width=950, height=850, bg='steelblue')
my_frame2 = Frame(my_notebook, width=950, height=850, bg='steelblue')

my_frame1.pack(fill='both', expand='yes')
my_frame2.pack(fill='both', expand='yes')

my_notebook.add(my_frame1, text='Insumos')
my_notebook.add(my_frame2, text='Fijos')
my_notebook.pack(fill='both', expand='yes')

### ETIQUETAS
cantidad = ttk.Label(my_frame1, text='Cantidad (Kg./Lit.)').place(x=115,y=25)
cantidad2 = ttk.Label(my_frame1, text='Cantidad (Kg./Lit.)').place(x=570,y=25)
cantidad2 = ttk.Label(my_frame1, text='Cantidad (Kg./Lit.)').place(x=1025,y=25)
precio = ttk.Label(my_frame1, text='Precio por unidad').place(x=230,y=25)
precio2 = ttk.Label(my_frame1, text='Precio por unidad').place(x=690,y=25)
precio2 = ttk.Label(my_frame1, text='Precio por unidad').place(x=1140,y=25)
total = ttk.Label(my_frame1, text='Costo total').place(x=345,y=25)
total = ttk.Label(my_frame1, text='Costo total').place(x=810,y=25)
total = ttk.Label(my_frame1, text='Costo total').place(x=1255,y=25)

precio_fijos = ttk.Label(my_frame2, text='Costo total').place(x=590, y=25)
obser_fijos = ttk.Label(my_frame2, text='Observación a realizar').place(x=730, y=25)

cantidad4 = ttk.Label(my_frame2, text='Cantidad (Kg./Lit.)').place(x=135,y=25)
precio4 = ttk.Label(my_frame2, text='Precio por unidad').place(x=250,y=25)
total4 = ttk.Label(my_frame2, text='Costo total').place(x=360,y=25)

### CHECKBOX
tarjeta_valor = tk.IntVar()
tarjeta = ttk.Checkbutton(text='Pago con tarjeta?', variable=tarjeta_valor, command=checkbox_clicked)
tarjeta.place(x=1025, y=600)


## COSTOS VARIOS - INSUMOS 1
label_carniceria = Label(my_frame1, text='CARNICERIA:')
label_carniceria.place(x=20, y=50)
label_pollo = Label(my_frame1, text='Pollo  ==> ')
label_pollo.place(x=20,y=75)
label_picada = ttk.Label(my_frame1, text='Picada  ==> ')
label_picada.place(x=20,y=100)
label_filet = ttk.Label(my_frame1, text='Filet  ==> ')
label_filet.place(x=20,y=125)
label_beef = ttk.Label(my_frame1, text='Roast Beef => ')
label_beef.place(x=20,y=150)
label_bondiola = ttk.Label(my_frame1, text='Bondiola  ==> ')
label_bondiola.place(x=20,y=175)
label_jamon = ttk.Label(my_frame1, text='Jamón  ==> ')
label_jamon.place(x=20,y=200)
label_nalga = ttk.Label(my_frame1, text='Nalga  ==> ')
label_nalga.place(x=20,y=225)
label_super = Label(my_frame1, text='SUPERMERCADO:')
label_super.place(x=20, y=250)
label_leche = Label(my_frame1, text='Leche ==> ')
label_leche.place(x=20, y=275)
label_manteca = Label(my_frame1, text='Manteca ==> ')
label_manteca.place(x=20, y=275)
label_aceite = Label(my_frame1, text='Aceite ==> ')
label_aceite.place(x=20, y=300)
label_harina = Label(my_frame1, text='Harina ==> ')
label_harina.place(x=20, y=325)
label_fideos = Label(my_frame1, text='Fideos ==> ')
label_fideos.place(x=20, y=350)
label_arroz = Label(my_frame1, text='Arroz ==> ')
label_arroz.place(x=20, y=375)
label_pur_tom = Label(my_frame1, text='Pure de Tom. => ')
label_pur_tom.place(x=20, y=400)
label_caldo = Label(my_frame1, text='Caldos ==> ')
label_caldo.place(x=20, y=425)
label_pana = Label(my_frame1, text='PANADERIA:')
label_pana.place(x=20, y=450)
label_pan = Label(my_frame1, text='Pan ==> ')
label_pan.place(x=20, y=475)
label_medial = Label(my_frame1, text='Medialunas ==> ')
label_medial.place(x=20, y=500)
label_gaseosa = Label(my_frame1, text='GASEOSAS:')
label_gaseosa.place(x=20, y=525)
label_sodag = Label(my_frame1, text='Soda grande ==> ')
label_sodag.place(x=20, y=550)
label_sodac = Label(my_frame1, text='Soda chica ==> ')
label_sodac.place(x=20, y=575)
label_cerve = Label(my_frame1, text='Cerveza ==> ')
label_cerve.place(x=20, y=600)
label_aguac = Label(my_frame1, text='Agua chica ==> ')
label_aguac.place(x=20, y=625)
label_aguag = Label(my_frame1, text='Agua grande => ')
label_aguag.place(x=20, y=650)
label_gaseosa = Label(my_frame1, text='Gaseosa ==> ')
label_gaseosa.place(x=20, y=675)
label_pan_rallado = Label(my_frame1, text='Pan Rallado =>')
label_pan_rallado.place(x=20, y=700)
label_verduleria = Label(my_frame1, text='VERDULERIA:')
label_verduleria.place(x=475, y=50)
label_cebolla = Label(my_frame1, text='Cebolla ==> ')
label_cebolla.place(x=475, y=75)
label_acelga = Label(my_frame1, text='Acelga ==> ')
label_acelga.place(x=475, y=100)
label_papa = Label(my_frame1, text='Papa ==> ')
label_papa.place(x=475, y=125)
label_cala = Label(my_frame1, text='Calabaza ==> ')
label_cala.place(x=475, y=150)
label_zapa = Label(my_frame1, text='Zapallito ==> ')
label_zapa.place(x=475, y=175)
label_beren = Label(my_frame1, text='Berenjenas ==> ')
label_beren.place(x=475, y=200)
label_puerro = Label(my_frame1, text='Puerro ==> ')
label_puerro.place(x=475, y=225)
label_morron = Label(my_frame1, text='Morrón ==> ')
label_morron.place(x=475, y=250)
label_tomate = Label(my_frame1, text='Tomate ==> ')
label_tomate.place(x=475, y=275)
label_verdeo = Label(my_frame1, text='Verdeo ==> ')
label_verdeo.place(x=475, y=300)
label_zana = Label(my_frame1, text='Zanahoria ==> ')
label_zana.place(x=475, y=325)
label_ajo = Label(my_frame1, text='Ajo ==> ')
label_ajo.place(x=475, y=350)
label_pechuga = Label(my_frame1, text='Pechuga ==> ')
label_pechuga.place(x=475, y=375)
label_batata = Label(my_frame1, text='Batata ==> ')
label_batata.place(x=475, y=400)
label_limon = Label(my_frame1, text='Limon ==> ')
label_limon.place(x=475, y=425)
label_dietetica = Label(my_frame1, text='DIETETICA:')
label_dietetica.place(x=475, y=450)
label_oregano = Label(my_frame1, text='Orégano ==> ')
label_oregano.place(x=475, y=475)
label_pimenton = Label(my_frame1, text='Pimentón ==> ')
label_pimenton.place(x=475, y=500)
label_pimienta = Label(my_frame1, text='Pimienta ==> ')
label_pimienta.place(x=475, y=525)
label_provenzal = Label(my_frame1, text='Provenzal ==> ')
label_provenzal.place(x=475, y=550)
label_nuez = Label(my_frame1, text='Nuez Moscada =>')
label_nuez.place(x=475, y=575)
label_laurel = Label(my_frame1, text='Laurel ==> ')
label_laurel.place(x=475, y=600)
label_otros = Label(my_frame1, text='OTROS:')
label_otros.place(x=475, y=625)
label_tapas = Label(my_frame1, text='Tapas de Emp.=>')
label_tapas.place(x=475, y=650)
label_noquis = Label(my_frame1, text='ñoquis ==>')
label_noquis.place(x=475, y=675)
label_huevos = Label(my_frame1, text='Huevos ==>')
label_huevos.place(x=475, y=700)
label_otros2 = Label(my_frame1, text='CAFETERIA:')
label_otros2.place(x=930, y=50)
label_te = Label(my_frame1, text='Té ==> ')
label_te.place(x=930, y=75)
label_azucar = Label(my_frame1, text='Azúcar ==> ')
label_azucar.place(x=930, y=100)
label_yerba = Label(my_frame1, text='Yerba ==> ')
label_yerba.place(x=930, y=125)
label_edulco = Label(my_frame1, text='Edulcorante ==> ')
label_edulco.place(x=930, y=150)
label_cafe_capsula = Label(my_frame1, text='Café Cápsula =>')
label_cafe_capsula.place(x=930, y=175)
label_cafe_molido = Label(my_frame1, text='Café Molido ==>')
label_cafe_molido.place(x=930, y=200)
label_cafe = Label(my_frame1, text='Café 1K ==>')
label_cafe.place(x=930, y=225)
label_vasos_cafe = Label(my_frame1, text='Vasos Café ==>')
label_vasos_cafe.place(x=930, y=250)

## COSTOS FIJOS
label_alquiler = Label(my_frame2, text='Alquiler  ==> ')
label_alquiler.place(x=475,y=60)
label_luz = Label(my_frame2, text='Luz  ==> ')
label_luz.place(x=475,y=85)
label_agua_servicio = Label(my_frame2, text='AYSA (Agua) =>')
label_agua_servicio.place(x=475,y=110)
label_telefono = Label(my_frame2, text='Teléfono  ==> ')
label_telefono.place(x=475,y=135)
label_abl = Label(my_frame2, text='ABL  ==> ')
label_abl.place(x=475,y=160)
label_diario = Label(my_frame2, text='Diario  ==> ')
label_diario.place(x=475,y=185)
label_seguro = Label(my_frame2, text='Fumigador  ==> ')
label_seguro.place(x=475,y=210)
label_detergente = Label(my_frame2, text='Seguro  ==> ')
label_detergente.place(x=475,y=235)
label_monotributo = Label(my_frame2, text='Monotributo  => ')
label_monotributo.place(x=475,y=260)
label_gas = Label(my_frame2, text='Gas  ==> ')
label_gas.place(x=475,y=285)
label_tarjeta = Label(my_frame2, text='Tarjeta  ==> ')
label_tarjeta.place(x=475,y=310)
label_iibb = Label(my_frame2, text='Ingresos Brutos=>')
label_iibb.place(x=475,y=335)
label_otros = Label(my_frame2, text='Otros =>')
label_otros.place(x=475,y=360)


# COSTOS FIJOS PARTE 2
label_descartables = Label(my_frame2, text='DESCARTABLES:')
label_descartables.place(x=20, y=25)
label_folex = Label(my_frame2, text='Flox  ==> ')
label_folex.place(x=20,y=60)
label_band = Label(my_frame2, text='Band. Cartón =>')
label_band.place(x=20,y=85)
label_sulfito = Label(my_frame2, text='Papel Sulfito =>')
label_sulfito.place(x=20,y=110)
label_film = Label(my_frame2, text='Papel Film ==>')
label_film.place(x=20,y=135)
label_servilletas = Label(my_frame2, text='Servilletas mesa =>')
label_servilletas.place(x=20,y=160)
label_servilletas2 = Label(my_frame2, text='Servilletas cocina =>')
label_servilletas2.place(x=20,y=185)
label_cubiertos = Label(my_frame2, text='Cubiertos ==>')
label_cubiertos.place(x=20,y=210)
label_bolsas = Label(my_frame2, text='Bolsas residuos =>')
label_bolsas.place(x=20,y=235)
label_bolsas2 = Label(my_frame2, text='Bolsas pedidos =>')
label_bolsas2.place(x=20,y=260)
label_remos = Label(my_frame2, text='Remos de café =>')
label_remos.place(x=20,y=285)
label_higienico = Label(my_frame2, text='Papel Higiénico =>')
label_higienico.place(x=20,y=310)
label_band2 = Label(my_frame2, text='Band. platos =>')
label_band2.place(x=20,y=335)
label_band3 = Label(my_frame2, text='Band. ensalada =>')
label_band3.place(x=20,y=360)
label_band4 = Label(my_frame2, text='Band. Puré =>')
label_band4.place(x=20,y=385)
label_band5 = Label(my_frame2, text='Band. con tapa =>')
label_band5.place(x=20,y=410)
label_alum = Label(my_frame2, text='Papel Aluminio =>')
label_alum.place(x=20,y=435)
label_limpieza = Label(my_frame2, text='LIMPIEZA:')
label_limpieza.place(x=20, y=460)
label_deter = Label(my_frame2, text='Detergente ==>')
label_deter.place(x=20, y=485)
label_lavandina = Label(my_frame2, text='Lavandina ==>')
label_lavandina.place(x=20, y=510)
label_perfume = Label(my_frame2, text='Perfume piso ==>')
label_perfume.place(x=20, y=535)
label_esponja = Label(my_frame2, text='Esponja ==>')
label_esponja.place(x=20, y=560)
label_jabon_polvo = Label(my_frame2, text='Jabón en polvo =>')
label_jabon_polvo.place(x=20, y=585)
label_desengrasante = Label(my_frame2, text='Desengrasante ==>')
label_desengrasante.place(x=20, y=610)
label_alcohol = Label(my_frame2, text='Alcohol ==>')
label_alcohol.place(x=20, y=635)


### CAJAS
## COSTOS VARIOS
caja_q_pollo = Entry(my_frame1)
caja_q_pollo.place(x=125, y=72,width=80)
caja_q_pollo.insert(tk.END,'')
caja_q_pollo.focus()
caja_p_pollo = Entry(my_frame1)
caja_p_pollo.place(x=235, y=72,width=80)
caja_p_pollo.insert(tk.END,'')
caja_t_pollo = Entry(my_frame1)
caja_t_pollo.place(x=345, y=72,width=80)
caja_t_pollo.insert(tk.END,'')

caja_q_picada = Entry(my_frame1)
caja_q_picada.place(x=125, y=98,width=80)
caja_q_picada.insert(tk.END,'')
caja_p_picada = Entry(my_frame1)
caja_p_picada.place(x=235, y=98,width=80)
caja_p_picada.insert(tk.END,'')
caja_t_picada = Entry(my_frame1)
caja_t_picada.place(x=345, y=98,width=80)
caja_t_picada.insert(tk.END,'')

caja_q_filet = Entry(my_frame1)
caja_q_filet.place(x=125, y=123,width=80)
caja_q_filet.insert(tk.END,'')
caja_p_filet = Entry(my_frame1)
caja_p_filet.place(x=235, y=123,width=80)
caja_p_filet.insert(tk.END,'')
caja_t_filet = Entry(my_frame1)
caja_t_filet.place(x=345, y=123,width=80)
caja_t_filet.insert(tk.END,'')

caja_q_beef = Entry(my_frame1)
caja_q_beef.place(x=125, y=148,width=80)
caja_q_beef.insert(tk.END,'')
caja_p_beef = Entry(my_frame1)
caja_p_beef.place(x=235, y=148,width=80)
caja_p_beef.insert(tk.END,'')
caja_t_beef = Entry(my_frame1)
caja_t_beef.place(x=345, y=148,width=80)
caja_t_beef.insert(tk.END,'')

caja_q_bondiola = Entry(my_frame1)
caja_q_bondiola.place(x=125, y=173,width=80)
caja_q_bondiola.insert(tk.END,'')
caja_p_bondiola = Entry(my_frame1)
caja_p_bondiola.place(x=235, y=173,width=80)
caja_p_bondiola.insert(tk.END,'')
caja_t_bondiola = Entry(my_frame1)
caja_t_bondiola.place(x=345, y=173,width=80)
caja_t_bondiola.insert(tk.END,'')

caja_q_jamon = Entry(my_frame1)
caja_q_jamon.place(x=125, y=198,width=80)
caja_q_jamon.insert(tk.END,'')
caja_p_jamon = Entry(my_frame1)
caja_p_jamon.place(x=235, y=198,width=80)
caja_p_jamon.insert(tk.END,'')
caja_t_jamon = Entry(my_frame1)
caja_t_jamon.place(x=345, y=198,width=80)
caja_t_jamon.insert(tk.END,'')

caja_q_nalga = Entry(my_frame1)
caja_q_nalga.place(x=125, y=223,width=80)
caja_q_nalga.insert(tk.END,'')
caja_p_nalga = Entry(my_frame1)
caja_p_nalga.place(x=235, y=223,width=80)
caja_p_nalga.insert(tk.END,'')
caja_t_nalga = Entry(my_frame1)
caja_t_nalga.place(x=345, y=223,width=80)
caja_t_nalga.insert(tk.END,'')

caja_q_manteca = Entry(my_frame1)
caja_q_manteca.place(x=125, y=273,width=80)
caja_q_manteca.insert(tk.END,'')
caja_p_manteca = Entry(my_frame1)
caja_p_manteca.place(x=235, y=273,width=80)
caja_p_manteca.insert(tk.END,'')
caja_t_manteca = Entry(my_frame1)
caja_t_manteca.place(x=345, y=273,width=80)
caja_t_manteca.insert(tk.END,'')

caja_q_aceite = Entry(my_frame1)
caja_q_aceite.place(x=125, y=298,width=80)
caja_q_aceite.insert(tk.END,'')
caja_p_aceite = Entry(my_frame1)
caja_p_aceite.place(x=235, y=298,width=80)
caja_p_aceite.insert(tk.END,'')
caja_t_aceite = Entry(my_frame1)
caja_t_aceite.place(x=345, y=298,width=80)
caja_t_aceite.insert(tk.END,'')

caja_q_harina = Entry(my_frame1)
caja_q_harina.place(x=125, y=323,width=80)
caja_q_harina.insert(tk.END,'')
caja_p_harina = Entry(my_frame1)
caja_p_harina.place(x=235, y=323,width=80)
caja_p_harina.insert(tk.END,'')
caja_t_harina = Entry(my_frame1)
caja_t_harina.place(x=345, y=323,width=80)
caja_t_harina.insert(tk.END,'')

caja_q_fideos = Entry(my_frame1)
caja_q_fideos.place(x=125, y=348,width=80)
caja_q_fideos.insert(tk.END,'')
caja_p_fideos = Entry(my_frame1)
caja_p_fideos.place(x=235, y=348,width=80)
caja_p_fideos.insert(tk.END,'')
caja_t_fideos = Entry(my_frame1)
caja_t_fideos.place(x=345, y=348,width=80)
caja_t_fideos.insert(tk.END,'')

caja_q_arroz = Entry(my_frame1)
caja_q_arroz.place(x=125, y=373,width=80)
caja_q_arroz.insert(tk.END,'')
caja_p_arroz = Entry(my_frame1)
caja_p_arroz.place(x=235, y=373,width=80)
caja_p_arroz.insert(tk.END,'')
caja_t_arroz = Entry(my_frame1)
caja_t_arroz.place(x=345, y=373,width=80)
caja_t_arroz.insert(tk.END,'')

caja_q_pure_tom = Entry(my_frame1)
caja_q_pure_tom.place(x=125, y=398,width=80)
caja_q_pure_tom.insert(tk.END,'')
caja_p_pure_tom = Entry(my_frame1)
caja_p_pure_tom.place(x=235, y=398,width=80)
caja_p_pure_tom.insert(tk.END,'')
caja_t_pure_tom = Entry(my_frame1)
caja_t_pure_tom.place(x=345, y=398,width=80)
caja_t_pure_tom.insert(tk.END,'')

caja_q_caldos = Entry(my_frame1)
caja_q_caldos.place(x=125, y=423,width=80)
caja_q_caldos.insert(tk.END,'')
caja_p_caldos = Entry(my_frame1)
caja_p_caldos.place(x=235, y=423,width=80)
caja_p_caldos.insert(tk.END,'')
caja_t_caldos = Entry(my_frame1)
caja_t_caldos.place(x=345, y=423,width=80)
caja_t_caldos.insert(tk.END,'')

caja_q_pan = Entry(my_frame1)
caja_q_pan.place(x=125, y=473,width=80)
caja_q_pan.insert(tk.END,'')
caja_p_pan = Entry(my_frame1)
caja_p_pan.place(x=235, y=473,width=80)
caja_p_pan.insert(tk.END,'')
caja_t_pan = Entry(my_frame1)
caja_t_pan.place(x=345, y=473,width=80)
caja_t_pan.insert(tk.END,'')

caja_q_medial = Entry(my_frame1)
caja_q_medial.place(x=125, y=498,width=80)
caja_q_medial.insert(tk.END,'')
caja_p_medial = Entry(my_frame1)
caja_p_medial.place(x=235, y=498,width=80)
caja_p_medial.insert(tk.END,'')
caja_t_medial = Entry(my_frame1)
caja_t_medial.place(x=345, y=498,width=80)
caja_t_medial.insert(tk.END,'')

caja_q_sodag = Entry(my_frame1)
caja_q_sodag.place(x=125, y=548,width=80)
caja_q_sodag.insert(tk.END,'')
caja_p_sodag = Entry(my_frame1)
caja_p_sodag.place(x=235, y=548,width=80)
caja_p_sodag.insert(tk.END,'')
caja_t_sodag = Entry(my_frame1)
caja_t_sodag.place(x=345, y=548,width=80)
caja_t_sodag.insert(tk.END,'')

caja_q_sodac = Entry(my_frame1)
caja_q_sodac.place(x=125, y=573,width=80)
caja_q_sodac.insert(tk.END,'')
caja_p_sodac = Entry(my_frame1)
caja_p_sodac.place(x=235, y=573,width=80)
caja_p_sodac.insert(tk.END,'')
caja_t_sodac = Entry(my_frame1)
caja_t_sodac.place(x=345, y=573,width=80)
caja_t_sodac.insert(tk.END,'')

caja_q_cerveza = Entry(my_frame1)
caja_q_cerveza.place(x=125, y=598,width=80)
caja_q_cerveza.insert(tk.END,'')
caja_p_cerveza = Entry(my_frame1)
caja_p_cerveza.place(x=235, y=598,width=80)
caja_p_cerveza.insert(tk.END,'')
caja_t_cerveza = Entry(my_frame1)
caja_t_cerveza.place(x=345, y=598,width=80)
caja_t_cerveza.insert(tk.END,'')

caja_q_aguac = Entry(my_frame1)
caja_q_aguac.place(x=125, y=623,width=80)
caja_q_aguac.insert(tk.END,'')
caja_p_aguac = Entry(my_frame1)
caja_p_aguac.place(x=235, y=623,width=80)
caja_p_aguac.insert(tk.END,'')
caja_t_aguac = Entry(my_frame1)
caja_t_aguac.place(x=345, y=623,width=80)
caja_t_aguac.insert(tk.END,'')

caja_q_aguag = Entry(my_frame1)
caja_q_aguag.place(x=125, y=648,width=80)
caja_q_aguag.insert(tk.END,'')
caja_p_aguag = Entry(my_frame1)
caja_p_aguag.place(x=235, y=648,width=80)
caja_p_aguag.insert(tk.END,'')
caja_t_aguag = Entry(my_frame1)
caja_t_aguag.place(x=345, y=648,width=80)
caja_t_aguag.insert(tk.END,'')

caja_q_gaseosa = Entry(my_frame1)
caja_q_gaseosa.place(x=125, y=673,width=80)
caja_q_gaseosa.insert(tk.END,'')
caja_p_gaseosa = Entry(my_frame1)
caja_p_gaseosa.place(x=235, y=673,width=80)
caja_p_gaseosa.insert(tk.END,'')
caja_t_gaseosa = Entry(my_frame1)
caja_t_gaseosa.place(x=345, y=673,width=80)
caja_t_gaseosa.insert(tk.END,'')

caja_q_pan_rallado = Entry(my_frame1)
caja_q_pan_rallado.place(x=125, y=698,width=80)
caja_q_pan_rallado.insert(tk.END,'')
caja_p_pan_rallado = Entry(my_frame1)
caja_p_pan_rallado.place(x=235, y=698,width=80)
caja_p_pan_rallado.insert(tk.END,'')
caja_t_pan_rallado = Entry(my_frame1)
caja_t_pan_rallado.place(x=345, y=698,width=80)
caja_t_pan_rallado.insert(tk.END,'')

caja_q_cebolla = Entry(my_frame1)
caja_q_cebolla.place(x=580, y=72,width=80)
caja_q_cebolla.insert(tk.END,'')
caja_p_cebolla = Entry(my_frame1)
caja_p_cebolla.place(x=690, y=72,width=80)
caja_p_cebolla.insert(tk.END,'')
caja_t_cebolla = Entry(my_frame1)
caja_t_cebolla.place(x=800, y=72,width=80)
caja_t_cebolla.insert(tk.END,'')

caja_q_acelga = Entry(my_frame1)
caja_q_acelga.place(x=580, y=98,width=80)
caja_q_acelga.insert(tk.END,'')
caja_p_acelga = Entry(my_frame1)
caja_p_acelga.place(x=690, y=98,width=80)
caja_p_acelga.insert(tk.END,'')
caja_t_acelga = Entry(my_frame1)
caja_t_acelga.place(x=800, y=98,width=80)
caja_t_acelga.insert(tk.END,'')

caja_q_papa = Entry(my_frame1)
caja_q_papa.place(x=580, y=123,width=80)
caja_q_papa.insert(tk.END,'')
caja_p_papa = Entry(my_frame1)
caja_p_papa.place(x=690, y=123,width=80)
caja_p_papa.insert(tk.END,'')
caja_t_papa = Entry(my_frame1)
caja_t_papa.place(x=800, y=123,width=80)
caja_t_papa.insert(tk.END,'')

caja_q_cala = Entry(my_frame1)
caja_q_cala.place(x=580, y=148,width=80)
caja_q_cala.insert(tk.END,'')
caja_p_cala = Entry(my_frame1)
caja_p_cala.place(x=690, y=148,width=80)
caja_p_cala.insert(tk.END,'')
caja_t_cala = Entry(my_frame1)
caja_t_cala.place(x=800, y=148,width=80)
caja_t_cala.insert(tk.END,'')

caja_q_zapa = Entry(my_frame1)
caja_q_zapa.place(x=580, y=173,width=80)
caja_q_zapa.insert(tk.END,'')
caja_p_zapa = Entry(my_frame1)
caja_p_zapa.place(x=690, y=173,width=80)
caja_p_zapa.insert(tk.END,'')
caja_t_zapa = Entry(my_frame1)
caja_t_zapa.place(x=800, y=173,width=80)
caja_t_zapa.insert(tk.END,'')

caja_q_beren = Entry(my_frame1)
caja_q_beren.place(x=580, y=198,width=80)
caja_q_beren.insert(tk.END,'')
caja_p_beren = Entry(my_frame1)
caja_p_beren.place(x=690, y=198,width=80)
caja_p_beren.insert(tk.END,'')
caja_t_beren = Entry(my_frame1)
caja_t_beren.place(x=800, y=198,width=80)
caja_t_beren.insert(tk.END,'')

caja_q_puerro = Entry(my_frame1)
caja_q_puerro.place(x=580, y=223,width=80)
caja_q_puerro.insert(tk.END,'')
caja_p_puerro = Entry(my_frame1)
caja_p_puerro.place(x=690, y=223,width=80)
caja_p_puerro.insert(tk.END,'')
caja_t_puerro = Entry(my_frame1)
caja_t_puerro.place(x=800, y=223,width=80)
caja_t_puerro.insert(tk.END,'')

caja_q_morron = Entry(my_frame1)
caja_q_morron.place(x=580, y=248,width=80)
caja_q_morron.insert(tk.END,'')
caja_p_morron = Entry(my_frame1)
caja_p_morron.place(x=690, y=248,width=80)
caja_p_morron.insert(tk.END,'')
caja_t_morron = Entry(my_frame1)
caja_t_morron.place(x=800, y=248,width=80)
caja_t_morron.insert(tk.END,'')

caja_q_tomate = Entry(my_frame1)
caja_q_tomate.place(x=580, y=273,width=80)
caja_q_tomate.insert(tk.END,'')
caja_p_tomate = Entry(my_frame1)
caja_p_tomate.place(x=690, y=273,width=80)
caja_p_tomate.insert(tk.END,'')
caja_t_tomate = Entry(my_frame1)
caja_t_tomate.place(x=800, y=273,width=80)
caja_t_tomate.insert(tk.END,'')

caja_q_verdeo = Entry(my_frame1)
caja_q_verdeo.place(x=580, y=298,width=80)
caja_q_verdeo.insert(tk.END,'')
caja_p_verdeo = Entry(my_frame1)
caja_p_verdeo.place(x=690, y=298,width=80)
caja_p_verdeo.insert(tk.END,'')
caja_t_verdeo = Entry(my_frame1)
caja_t_verdeo.place(x=800, y=298,width=80)
caja_t_verdeo.insert(tk.END,'')

caja_q_zana = Entry(my_frame1)
caja_q_zana.place(x=580, y=323,width=80)
caja_q_zana.insert(tk.END,'')
caja_p_zana = Entry(my_frame1)
caja_p_zana.place(x=690, y=323,width=80)
caja_p_zana.insert(tk.END,'')
caja_t_zana = Entry(my_frame1)
caja_t_zana.place(x=800, y=323,width=80)
caja_t_zana.insert(tk.END,'')

caja_q_ajo = Entry(my_frame1)
caja_q_ajo.place(x=580, y=348,width=80)
caja_q_ajo.insert(tk.END,'')
caja_p_ajo = Entry(my_frame1)
caja_p_ajo.place(x=690, y=348,width=80)
caja_p_ajo.insert(tk.END,'')
caja_t_ajo = Entry(my_frame1)
caja_t_ajo.place(x=800, y=348,width=80)
caja_t_ajo.insert(tk.END,'')

caja_q_pechuga = Entry(my_frame1)
caja_q_pechuga.place(x=580, y=373,width=80)
caja_q_pechuga.insert(tk.END,'')
caja_p_pechuga = Entry(my_frame1)
caja_p_pechuga.place(x=690, y=373,width=80)
caja_p_pechuga.insert(tk.END,'')
caja_t_pechuga = Entry(my_frame1)
caja_t_pechuga.place(x=800, y=373,width=80)
caja_t_pechuga.insert(tk.END,'')

caja_q_batata = Entry(my_frame1)
caja_q_batata.place(x=580, y=398,width=80)
caja_q_batata.insert(tk.END,'')
caja_p_batata = Entry(my_frame1)
caja_p_batata.place(x=690, y=398,width=80)
caja_p_batata.insert(tk.END,'')
caja_t_batata = Entry(my_frame1)
caja_t_batata.place(x=800, y=398,width=80)
caja_t_batata.insert(tk.END,'')

caja_q_limon = Entry(my_frame1)
caja_q_limon.place(x=580, y=423,width=80)
caja_q_limon.insert(tk.END,'')
caja_p_limon = Entry(my_frame1)
caja_p_limon.place(x=690, y=423,width=80)
caja_p_limon.insert(tk.END,'')
caja_t_limon = Entry(my_frame1)
caja_t_limon.place(x=800, y=423,width=80)
caja_t_limon.insert(tk.END,'')

caja_q_oregano = Entry(my_frame1)
caja_q_oregano.place(x=580, y=473,width=80)
caja_q_oregano.insert(tk.END,'')
caja_p_oregano = Entry(my_frame1)
caja_p_oregano.place(x=690, y=473,width=80)
caja_p_oregano.insert(tk.END,'')
caja_t_oregano = Entry(my_frame1)
caja_t_oregano.place(x=800, y=473,width=80)
caja_t_oregano.insert(tk.END,'')

caja_q_pimenton = Entry(my_frame1)
caja_q_pimenton.place(x=580, y=498,width=80)
caja_q_pimenton.insert(tk.END,'')
caja_p_pimenton = Entry(my_frame1)
caja_p_pimenton.place(x=690, y=498,width=80)
caja_p_pimenton.insert(tk.END,'')
caja_t_pimenton = Entry(my_frame1)
caja_t_pimenton.place(x=800, y=498,width=80)
caja_t_pimenton.insert(tk.END,'')

caja_q_pimienta = Entry(my_frame1)
caja_q_pimienta.place(x=580, y=523,width=80)
caja_q_pimienta.insert(tk.END,'')
caja_p_pimienta = Entry(my_frame1)
caja_p_pimienta.place(x=690, y=523,width=80)
caja_p_pimienta.insert(tk.END,'')
caja_t_pimienta = Entry(my_frame1)
caja_t_pimienta.place(x=800, y=523,width=80)
caja_t_pimienta.insert(tk.END,'')

caja_q_provenzal = Entry(my_frame1)
caja_q_provenzal.place(x=580, y=548,width=80)
caja_q_provenzal.insert(tk.END,'')
caja_p_provenzal = Entry(my_frame1)
caja_p_provenzal.place(x=690, y=548,width=80)
caja_p_provenzal.insert(tk.END,'')
caja_t_provenzal = Entry(my_frame1)
caja_t_provenzal.place(x=800, y=548,width=80)
caja_t_provenzal.insert(tk.END,'')

caja_q_nuez = Entry(my_frame1)
caja_q_nuez.place(x=580, y=573,width=80)
caja_q_nuez.insert(tk.END,'')
caja_p_nuez = Entry(my_frame1)
caja_p_nuez.place(x=690, y=573,width=80)
caja_p_nuez.insert(tk.END,'')
caja_t_nuez = Entry(my_frame1)
caja_t_nuez.place(x=800, y=573,width=80)
caja_t_nuez.insert(tk.END,'')

caja_q_laurel = Entry(my_frame1)
caja_q_laurel.place(x=580, y=598,width=80)
caja_q_laurel.insert(tk.END,'')
caja_p_laurel = Entry(my_frame1)
caja_p_laurel.place(x=690, y=598,width=80)
caja_p_laurel.insert(tk.END,'')
caja_t_laurel = Entry(my_frame1)
caja_t_laurel.place(x=800, y=598,width=80)
caja_t_laurel.insert(tk.END,'')

caja_q_tapas = Entry(my_frame1)
caja_q_tapas.place(x=580, y=648,width=80)
caja_q_tapas.insert(tk.END,'')
caja_p_tapas = Entry(my_frame1)
caja_p_tapas.place(x=690, y=648,width=80)
caja_p_tapas.insert(tk.END,'')
caja_t_tapas = Entry(my_frame1)
caja_t_tapas.place(x=800, y=648,width=80)
caja_t_tapas.insert(tk.END,'')

caja_q_noquiz = Entry(my_frame1)
caja_q_noquiz.place(x=580, y=673,width=80)
caja_q_noquiz.insert(tk.END,'')
caja_p_noquiz = Entry(my_frame1)
caja_p_noquiz.place(x=690, y=673,width=80)
caja_p_noquiz.insert(tk.END,'')
caja_t_noquiz = Entry(my_frame1)
caja_t_noquiz.place(x=800, y=673,width=80)
caja_t_noquiz.insert(tk.END,'')

caja_q_huevos = Entry(my_frame1)
caja_q_huevos.place(x=580, y=698,width=80)
caja_q_huevos.insert(tk.END,'')
caja_p_huevos = Entry(my_frame1)
caja_p_huevos.place(x=690, y=698,width=80)
caja_p_huevos.insert(tk.END,'')
caja_t_huevos = Entry(my_frame1)
caja_t_huevos.place(x=800, y=698,width=80)
caja_t_huevos.insert(tk.END,'')

caja_q_te = Entry(my_frame1)
caja_q_te.place(x=1035, y=75,width=80)
caja_q_te.insert(tk.END,'')
caja_p_te = Entry(my_frame1)
caja_p_te.place(x=1145, y=75,width=80)
caja_p_te.insert(tk.END,'')
caja_t_te = Entry(my_frame1)
caja_t_te.place(x=1250, y=75,width=80)
caja_t_te.insert(tk.END,'')

caja_q_azucar = Entry(my_frame1)
caja_q_azucar.place(x=1035, y=100,width=80)
caja_q_azucar.insert(tk.END,'')
caja_p_azucar = Entry(my_frame1)
caja_p_azucar.place(x=1145, y=100,width=80)
caja_p_azucar.insert(tk.END,'')
caja_t_azucar = Entry(my_frame1)
caja_t_azucar.place(x=1250, y=100,width=80)
caja_t_azucar.insert(tk.END,'')

caja_q_yerba = Entry(my_frame1)
caja_q_yerba.place(x=1035, y=125,width=80)
caja_q_yerba.insert(tk.END,'')
caja_p_yerba = Entry(my_frame1)
caja_p_yerba.place(x=1145, y=125,width=80)
caja_p_yerba.insert(tk.END,'')
caja_t_yerba = Entry(my_frame1)
caja_t_yerba.place(x=1250, y=125,width=80)
caja_t_yerba.insert(tk.END,'')

caja_q_edulco = Entry(my_frame1)
caja_q_edulco.place(x=1035, y=150,width=80)
caja_q_edulco.insert(tk.END,'')
caja_p_edulco = Entry(my_frame1)
caja_p_edulco.place(x=1145, y=150,width=80)
caja_p_edulco.insert(tk.END,'')
caja_t_edulco = Entry(my_frame1)
caja_t_edulco.place(x=1250, y=150,width=80)
caja_t_edulco.insert(tk.END,'')

caja_q_capsula = Entry(my_frame1)
caja_q_capsula.place(x=1035, y=175,width=80)
caja_q_capsula.insert(tk.END,'')
caja_p_capsula = Entry(my_frame1)
caja_p_capsula.place(x=1145, y=175,width=80)
caja_p_capsula.insert(tk.END,'')
caja_t_capsula = Entry(my_frame1)
caja_t_capsula.place(x=1250, y=175,width=80)
caja_t_capsula.insert(tk.END,'')

caja_q_molido = Entry(my_frame1)
caja_q_molido.place(x=1035, y=200,width=80)
caja_q_molido.insert(tk.END,'')
caja_p_molido = Entry(my_frame1)
caja_p_molido.place(x=1145, y=200,width=80)
caja_p_molido.insert(tk.END,'')
caja_t_molido = Entry(my_frame1)
caja_t_molido.place(x=1250, y=200,width=80)
caja_t_molido.insert(tk.END,'')

caja_q_cafe_kilo = Entry(my_frame1)
caja_q_cafe_kilo.place(x=1035, y=225,width=80)
caja_q_cafe_kilo.insert(tk.END,'')
caja_p_cafe_kilo = Entry(my_frame1)
caja_p_cafe_kilo.place(x=1145, y=225,width=80)
caja_p_cafe_kilo.insert(tk.END,'')
caja_t_cafe_kilo = Entry(my_frame1)
caja_t_cafe_kilo.place(x=1250, y=225,width=80)
caja_t_cafe_kilo.insert(tk.END,'')

caja_q_vasos_cafe = Entry(my_frame1)
caja_q_vasos_cafe.place(x=1035, y=250,width=80)
caja_q_vasos_cafe.insert(tk.END,'')
caja_p_vasos_cafe = Entry(my_frame1)
caja_p_vasos_cafe.place(x=1145, y=250,width=80)
caja_p_vasos_cafe.insert(tk.END,'')
caja_t_vasos_cafe = Entry(my_frame1)
caja_t_vasos_cafe.place(x=1250, y=250,width=80)
caja_t_vasos_cafe.insert(tk.END,'')


caja_total = ttk.Entry()
caja_total.place(x=1025, y=640, width=160)
caja_total.insert(tk.END,'')

## COSTOS FIJOS
caja_p_alquiler = Entry(my_frame2)
caja_p_alquiler.place(x=580, y=60,width=80)
caja_p_alquiler.insert(tk.END,'')
caja_obs_alquiler = Entry(my_frame2)
caja_obs_alquiler.place(x=690, y=60,width=200)
caja_obs_alquiler.insert(tk.END,'')

caja_p_luz = Entry(my_frame2)
caja_p_luz.place(x=580, y=85,width=80)
caja_p_luz.insert(tk.END,'')
caja_obs_luz = Entry(my_frame2)
caja_obs_luz.place(x=690, y=85,width=200)
caja_obs_luz.insert(tk.END,'')

caja_p_agua_servicio = Entry(my_frame2)
caja_p_agua_servicio.place(x=580, y=110,width=80)
caja_p_agua_servicio.insert(tk.END,'')
caja_obs_agua = Entry(my_frame2)
caja_obs_agua.place(x=690, y=110,width=200)
caja_obs_agua.insert(tk.END,'')

caja_p_telefono = Entry(my_frame2)
caja_p_telefono.place(x=580, y=135,width=80)
caja_p_telefono.insert(tk.END,'')
caja_obs_telefono = Entry(my_frame2)
caja_obs_telefono.place(x=690, y=135,width=200)
caja_obs_telefono.insert(tk.END,'')

caja_p_abl = Entry(my_frame2)
caja_p_abl.place(x=580, y=160,width=80)
caja_p_abl.insert(tk.END,'')
caja_obs_abl = Entry(my_frame2)
caja_obs_abl.place(x=690, y=160,width=200)
caja_obs_abl.insert(tk.END,'')

caja_p_diario = Entry(my_frame2)
caja_p_diario.place(x=580, y=185,width=80)
caja_p_diario.insert(tk.END,'')
caja_obs_diario = Entry(my_frame2)
caja_obs_diario.place(x=690, y=185,width=200)
caja_obs_diario.insert(tk.END,'')

caja_p_fumigador = Entry(my_frame2)
caja_p_fumigador.place(x=580, y=210,width=80)
caja_p_fumigador.insert(tk.END,'')
caja_obs_fumigador = Entry(my_frame2)
caja_obs_fumigador.place(x=690, y=210,width=200)
caja_obs_fumigador.insert(tk.END,'')

caja_p_detergente = Entry(my_frame2)
caja_p_detergente.place(x=580, y=235,width=80)
caja_p_detergente.insert(tk.END,'')
caja_obs_detergente = Entry(my_frame2)
caja_obs_detergente.place(x=690, y=235,width=200)
caja_obs_detergente.insert(tk.END,'')

caja_p_monotributo = Entry(my_frame2)
caja_p_monotributo.place(x=580, y=260,width=80)
caja_p_monotributo.insert(tk.END,'')
caja_obs_monotributo = Entry(my_frame2)
caja_obs_monotributo.place(x=690, y=260,width=200)
caja_obs_monotributo.insert(tk.END,'')

caja_p_gas = Entry(my_frame2)
caja_p_gas.place(x=580, y=285,width=80)
caja_p_gas.insert(tk.END,'')
caja_obs_gas = Entry(my_frame2)
caja_obs_gas.place(x=690, y=285,width=200)
caja_obs_gas.insert(tk.END,'')

caja_p_tarjeta = Entry(my_frame2)
caja_p_tarjeta.place(x=580, y=310,width=80)
caja_p_tarjeta.insert(tk.END,'')
caja_obs_tarjeta = Entry(my_frame2)
caja_obs_tarjeta.place(x=690, y=310,width=200)
caja_obs_tarjeta.insert(tk.END,'')

caja_p_iibb = Entry(my_frame2)
caja_p_iibb.place(x=580, y=335,width=80)
caja_p_iibb.insert(tk.END,'')
caja_obs_iibb = Entry(my_frame2)
caja_obs_iibb.place(x=690, y=335,width=200)
caja_obs_iibb.insert(tk.END,'')

caja_p_otros = Entry(my_frame2)
caja_p_otros.place(x=580, y=360,width=80)
caja_p_otros.insert(tk.END,'')
caja_obs_otros = Entry(my_frame2)
caja_obs_otros.place(x=690, y=360,width=200)
caja_obs_otros.insert(tk.END,'')


# CAJAS COSTOS FIJOS 2
caja_q_flox = Entry(my_frame2)
caja_q_flox.place(x=140, y=60,width=80)
caja_q_flox.insert(tk.END,'')
caja_p_flox = Entry(my_frame2)
caja_p_flox.place(x=255, y=60,width=80)
caja_p_flox.insert(tk.END,'')
caja_t_flox = Entry(my_frame2)
caja_t_flox.place(x=360, y=60,width=80)
caja_t_flox.insert(tk.END,'')

caja_q_carton = Entry(my_frame2)
caja_q_carton.place(x=140, y=85,width=80)
caja_q_carton.insert(tk.END,'')
caja_p_carton = Entry(my_frame2)
caja_p_carton.place(x=255, y=85,width=80)
caja_p_carton.insert(tk.END,'')
caja_t_carton = Entry(my_frame2)
caja_t_carton.place(x=360, y=85,width=80)
caja_t_carton.insert(tk.END,'')

caja_q_sulfito = Entry(my_frame2)
caja_q_sulfito.place(x=140, y=110,width=80)
caja_q_sulfito.insert(tk.END,'')
caja_p_sulfito = Entry(my_frame2)
caja_p_sulfito.place(x=255, y=110,width=80)
caja_p_sulfito.insert(tk.END,'')
caja_t_sulfito = Entry(my_frame2)
caja_t_sulfito.place(x=360, y=110,width=80)
caja_t_sulfito.insert(tk.END,'')

caja_q_film = Entry(my_frame2)
caja_q_film.place(x=140, y=135,width=80)
caja_q_film.insert(tk.END,'')
caja_p_film = Entry(my_frame2)
caja_p_film.place(x=255, y=135,width=80)
caja_p_film.insert(tk.END,'')
caja_t_film = Entry(my_frame2)
caja_t_film.place(x=360, y=135,width=80)
caja_t_film.insert(tk.END,'')

caja_q_serv_mesa = Entry(my_frame2)
caja_q_serv_mesa.place(x=140, y=160,width=80)
caja_q_serv_mesa.insert(tk.END,'')
caja_p_serv_mesa = Entry(my_frame2)
caja_p_serv_mesa.place(x=255, y=160,width=80)
caja_p_serv_mesa.insert(tk.END,'')
caja_t_serv_mesa = Entry(my_frame2)
caja_t_serv_mesa.place(x=360, y=160,width=80)
caja_t_serv_mesa.insert(tk.END,'')

caja_q_serv_cocina = Entry(my_frame2)
caja_q_serv_cocina.place(x=140, y=185,width=80)
caja_q_serv_cocina.insert(tk.END,'')
caja_p_serv_cocina = Entry(my_frame2)
caja_p_serv_cocina.place(x=255, y=185,width=80)
caja_p_serv_cocina.insert(tk.END,'')
caja_t_serv_cocina = Entry(my_frame2)
caja_t_serv_cocina.place(x=360, y=185,width=80)
caja_t_serv_cocina.insert(tk.END,'')

caja_q_cubiertos = Entry(my_frame2)
caja_q_cubiertos.place(x=140, y=210,width=80)
caja_q_cubiertos.insert(tk.END,'')
caja_p_cubiertos = Entry(my_frame2)
caja_p_cubiertos.place(x=255, y=210,width=80)
caja_p_cubiertos.insert(tk.END,'')
caja_t_cubiertos = Entry(my_frame2)
caja_t_cubiertos.place(x=360, y=210,width=80)
caja_t_cubiertos.insert(tk.END,'')

caja_q_bolsas_residuos = Entry(my_frame2)
caja_q_bolsas_residuos.place(x=140, y=235,width=80)
caja_q_bolsas_residuos.insert(tk.END,'')
caja_p_bolsas_residuos = Entry(my_frame2)
caja_p_bolsas_residuos.place(x=255, y=235,width=80)
caja_p_bolsas_residuos.insert(tk.END,'')
caja_t_bolsas_residuos = Entry(my_frame2)
caja_t_bolsas_residuos.place(x=360, y=235,width=80)
caja_t_bolsas_residuos.insert(tk.END,'')

caja_q_bolsas_pedidos = Entry(my_frame2)
caja_q_bolsas_pedidos.place(x=140, y=260,width=80)
caja_q_bolsas_pedidos.insert(tk.END,'')
caja_p_bolsas_pedidos = Entry(my_frame2)
caja_p_bolsas_pedidos.place(x=255, y=260,width=80)
caja_p_bolsas_pedidos.insert(tk.END,'')
caja_t_bolsas_pedidos = Entry(my_frame2)
caja_t_bolsas_pedidos.place(x=360, y=260,width=80)
caja_t_bolsas_pedidos.insert(tk.END,'')

caja_q_remos = Entry(my_frame2)
caja_q_remos.place(x=140, y=285,width=80)
caja_q_remos.insert(tk.END,'')
caja_p_remos = Entry(my_frame2)
caja_p_remos.place(x=255, y=285,width=80)
caja_p_remos.insert(tk.END,'')
caja_t_remos = Entry(my_frame2)
caja_t_remos.place(x=360, y=285,width=80)
caja_t_remos.insert(tk.END,'')

caja_q_higienico = Entry(my_frame2)
caja_q_higienico.place(x=140, y=310,width=80)
caja_q_higienico.insert(tk.END,'')
caja_p_higienico = Entry(my_frame2)
caja_p_higienico.place(x=255, y=310,width=80)
caja_p_higienico.insert(tk.END,'')
caja_t_higienico = Entry(my_frame2)
caja_t_higienico.place(x=360, y=310,width=80)
caja_t_higienico.insert(tk.END,'')

caja_q_platos = Entry(my_frame2)
caja_q_platos.place(x=140, y=335,width=80)
caja_q_platos.insert(tk.END,'')
caja_p_platos = Entry(my_frame2)
caja_p_platos.place(x=255, y=335,width=80)
caja_p_platos.insert(tk.END,'')
caja_t_platos = Entry(my_frame2)
caja_t_platos.place(x=360, y=335,width=80)
caja_t_platos.insert(tk.END,'')

caja_q_band_ensa = Entry(my_frame2)
caja_q_band_ensa.place(x=140, y=360,width=80)
caja_q_band_ensa.insert(tk.END,'')
caja_p_band_ensa = Entry(my_frame2)
caja_p_band_ensa.place(x=255, y=360,width=80)
caja_p_band_ensa.insert(tk.END,'')
caja_t_band_ensa = Entry(my_frame2)
caja_t_band_ensa.place(x=360, y=360,width=80)
caja_t_band_ensa.insert(tk.END,'')

caja_q_band_pure = Entry(my_frame2)
caja_q_band_pure.place(x=140, y=385,width=80)
caja_q_band_pure.insert(tk.END,'')
caja_p_band_pure = Entry(my_frame2)
caja_p_band_pure.place(x=255, y=385,width=80)
caja_p_band_pure.insert(tk.END,'')
caja_t_band_pure = Entry(my_frame2)
caja_t_band_pure.place(x=360, y=385,width=80)
caja_t_band_pure.insert(tk.END,'')

caja_q_band_tapa = Entry(my_frame2)
caja_q_band_tapa.place(x=140, y=410,width=80)
caja_q_band_tapa.insert(tk.END,'')
caja_p_band_tapa = Entry(my_frame2)
caja_p_band_tapa.place(x=255, y=410,width=80)
caja_p_band_tapa.insert(tk.END,'')
caja_t_band_tapa = Entry(my_frame2)
caja_t_band_tapa.place(x=360, y=410,width=80)
caja_t_band_tapa.insert(tk.END,'')

caja_q_alum = Entry(my_frame2)
caja_q_alum.place(x=140, y=435,width=80)
caja_q_alum.insert(tk.END,'')
caja_p_alum = Entry(my_frame2)
caja_p_alum.place(x=255, y=435,width=80)
caja_p_alum.insert(tk.END,'')
caja_t_alum = Entry(my_frame2)
caja_t_alum.place(x=360, y=435,width=80)
caja_t_alum.insert(tk.END,'')

caja_q_detergente = Entry(my_frame2)
caja_q_detergente.place(x=140, y=480,width=80)
caja_q_detergente.insert(tk.END,'')
caja_p_detergente = Entry(my_frame2)
caja_p_detergente.place(x=255, y=480,width=80)
caja_p_detergente.insert(tk.END,'')
caja_t_detergente = Entry(my_frame2)
caja_t_detergente.place(x=360, y=480,width=80)
caja_t_detergente.insert(tk.END,'')

caja_q_lavan = Entry(my_frame2)
caja_q_lavan.place(x=140, y=510,width=80)
caja_q_lavan.insert(tk.END,'')
caja_p_lavan = Entry(my_frame2)
caja_p_lavan.place(x=255, y=510,width=80)
caja_p_lavan.insert(tk.END,'')
caja_t_lavan = Entry(my_frame2)
caja_t_lavan.place(x=360, y=510,width=80)
caja_t_lavan.insert(tk.END,'')

caja_q_perfume = Entry(my_frame2)
caja_q_perfume.place(x=140, y=535,width=80)
caja_q_perfume.insert(tk.END,'')
caja_p_perfume = Entry(my_frame2)
caja_p_perfume.place(x=255, y=535,width=80)
caja_p_perfume.insert(tk.END,'')
caja_t_perfume = Entry(my_frame2)
caja_t_perfume.place(x=360, y=535,width=80)
caja_t_perfume.insert(tk.END,'')

caja_q_esponja = Entry(my_frame2)
caja_q_esponja.place(x=140, y=560,width=80)
caja_q_esponja.insert(tk.END,'')
caja_p_esponja = Entry(my_frame2)
caja_p_esponja.place(x=255, y=560,width=80)
caja_p_esponja.insert(tk.END,'')
caja_t_esponja = Entry(my_frame2)
caja_t_esponja.place(x=360, y=560,width=80)
caja_t_esponja.insert(tk.END,'')

caja_q_jabon = Entry(my_frame2)
caja_q_jabon.place(x=140, y=585,width=80)
caja_q_jabon.insert(tk.END,'')
caja_p_jabon = Entry(my_frame2)
caja_p_jabon.place(x=255, y=585,width=80)
caja_p_jabon.insert(tk.END,'')
caja_t_jabon = Entry(my_frame2)
caja_t_jabon.place(x=360, y=585,width=80)
caja_t_jabon.insert(tk.END,'')

caja_q_desengrasante = Entry(my_frame2)
caja_q_desengrasante.place(x=140, y=610,width=80)
caja_q_desengrasante.insert(tk.END,'')
caja_p_desengrasante = Entry(my_frame2)
caja_p_desengrasante.place(x=255, y=610,width=80)
caja_p_desengrasante.insert(tk.END,'')
caja_t_desengrasante = Entry(my_frame2)
caja_t_desengrasante.place(x=360, y=610,width=80)
caja_t_desengrasante.insert(tk.END,'')

caja_q_alcohol = Entry(my_frame2)
caja_q_alcohol.place(x=140, y=635,width=80)
caja_q_alcohol.insert(tk.END,'')
caja_p_alcohol = Entry(my_frame2)
caja_p_alcohol.place(x=255, y=635,width=80)
caja_p_alcohol.insert(tk.END,'')
caja_t_alcohol = Entry(my_frame2)
caja_t_alcohol.place(x=360, y=635,width=80)
caja_t_alcohol.insert(tk.END,'')

cajas = [caja_q_pollo, caja_p_pollo, caja_t_pollo, caja_q_picada, caja_p_picada, caja_t_picada,caja_q_filet,
         caja_p_filet, caja_t_filet, caja_q_beef, caja_p_beef, caja_t_beef, caja_q_bondiola, caja_p_bondiola,
         caja_t_bondiola, caja_q_jamon, caja_p_jamon, caja_t_jamon, caja_q_nalga, caja_p_nalga, caja_t_nalga,
         caja_q_manteca, caja_p_manteca, caja_t_manteca, caja_q_aceite, caja_p_aceite, caja_t_aceite, caja_q_harina,
         caja_p_harina, caja_t_harina, caja_q_fideos, caja_p_fideos, caja_t_fideos, caja_q_arroz, caja_p_arroz,
         caja_t_arroz, caja_q_pure_tom, caja_p_pure_tom, caja_t_pure_tom, caja_q_caldos, caja_p_caldos,
         caja_t_caldos, caja_q_pan, caja_p_pan, caja_t_pan, caja_q_medial, caja_p_medial, caja_t_medial,
         caja_q_sodag, caja_p_sodag, caja_t_sodag, caja_q_sodac, caja_p_sodac, caja_t_sodac, caja_q_cerveza,
         caja_p_cerveza, caja_t_cerveza, caja_q_aguac, caja_p_aguac, caja_t_aguac, caja_q_aguag, caja_p_aguag,
         caja_t_aguag, caja_q_gaseosa, caja_p_aguag, caja_t_aguag, caja_q_gaseosa, caja_p_gaseosa, caja_t_gaseosa,
         caja_q_pan_rallado, caja_p_pan_rallado, caja_t_pan_rallado, caja_q_cebolla, caja_p_cebolla, caja_t_cebolla, caja_t_cala,
         caja_q_acelga, caja_p_acelga, caja_t_acelga, caja_q_papa, caja_p_papa, caja_t_papa, caja_q_cala, caja_p_cala,
         caja_t_zapa, caja_q_zapa, caja_p_zapa, caja_q_beren, caja_p_beren, caja_t_beren, caja_q_puerro, caja_p_puerro, caja_t_puerro,
         caja_q_morron, caja_p_morron, caja_t_morron, caja_q_tomate, caja_p_tomate, caja_t_tomate, caja_q_verdeo,
         caja_p_verdeo, caja_t_verdeo, caja_q_zana, caja_p_zana, caja_t_zana, caja_q_ajo, caja_p_ajo, caja_t_ajo,
         caja_q_pechuga, caja_p_pechuga, caja_t_pechuga, caja_q_batata, caja_p_batata, caja_t_batata, caja_q_limon,
         caja_p_limon, caja_t_limon, caja_q_oregano, caja_p_oregano, caja_t_oregano, caja_q_pimenton, caja_p_pimenton,
         caja_t_pimenton, caja_q_pimienta, caja_p_pimienta, caja_t_pimienta, caja_q_provenzal, caja_p_provenzal,
         caja_t_provenzal, caja_q_nuez, caja_p_nuez, caja_t_nuez, caja_q_laurel, caja_p_laurel, caja_t_laurel,
         caja_q_tapas, caja_p_tapas, caja_t_tapas, caja_q_noquiz, caja_p_noquiz, caja_t_noquiz, caja_q_huevos,
         caja_p_huevos, caja_t_huevos, caja_q_te, caja_p_te, caja_t_te, caja_q_azucar, caja_p_azucar, caja_t_azucar,
         caja_q_yerba, caja_p_yerba, caja_t_yerba, caja_q_edulco, caja_p_edulco, caja_t_edulco, caja_q_capsula,
         caja_p_capsula, caja_t_capsula, caja_q_molido, caja_p_molido, caja_t_molido, caja_q_cafe_kilo, caja_p_cafe_kilo,
         caja_t_cafe_kilo, caja_q_vasos_cafe, caja_p_vasos_cafe, caja_t_vasos_cafe, caja_p_alquiler,
         caja_obs_alquiler, caja_p_luz, caja_obs_luz, caja_p_agua_servicio, caja_obs_agua, caja_p_telefono,
         caja_obs_telefono, caja_p_abl, caja_obs_abl, caja_p_diario, caja_obs_diario, caja_p_fumigador,
         caja_obs_fumigador, caja_p_detergente, caja_obs_detergente, caja_p_monotributo, caja_obs_monotributo,
         caja_p_gas, caja_obs_gas, caja_p_tarjeta, caja_obs_tarjeta, caja_p_iibb, caja_obs_iibb, caja_p_otros, caja_obs_otros,
         caja_q_flox, caja_p_flox, caja_t_flox, caja_q_carton, caja_p_carton, caja_t_carton, caja_q_sulfito, caja_p_sulfito,
         caja_t_sulfito, caja_q_film, caja_p_film, caja_t_film, caja_q_serv_mesa, caja_p_serv_mesa, caja_t_serv_mesa,
         caja_q_serv_cocina, caja_p_serv_mesa, caja_t_serv_mesa, caja_q_serv_cocina, caja_p_serv_cocina,
         caja_t_serv_cocina, caja_q_cubiertos, caja_p_cubiertos, caja_t_cubiertos, caja_q_bolsas_residuos,
         caja_p_bolsas_residuos, caja_t_bolsas_residuos, caja_q_bolsas_pedidos, caja_p_bolsas_pedidos,
         caja_t_bolsas_pedidos, caja_q_remos, caja_p_remos, caja_t_remos, caja_q_higienico, caja_p_higienico,
         caja_t_higienico, caja_q_platos, caja_p_platos, caja_t_platos, caja_q_band_ensa, caja_p_band_ensa,
         caja_t_band_ensa, caja_q_band_pure, caja_p_band_pure, caja_t_band_pure, caja_q_band_tapa, caja_p_band_pure,
         caja_t_band_pure, caja_q_band_tapa, caja_p_band_tapa, caja_t_band_tapa, caja_q_alum, caja_p_alum,
         caja_t_alum, caja_q_detergente, caja_p_detergente, caja_t_detergente, caja_q_lavan, caja_p_lavan,
         caja_t_lavan, caja_q_perfume, caja_p_perfume, caja_t_perfume, caja_q_esponja, caja_p_esponja, caja_t_esponja,
         caja_q_jabon, caja_p_jabon, caja_t_jabon, caja_q_desengrasante, caja_t_desengrasante, caja_p_desengrasante,
         caja_q_alcohol, caja_p_alcohol, caja_t_alcohol]


## BOTONES - MESSAGE BOX
costos = ttk.Button(text='Totalizar costos',command=mult)
costos.place(x=1025,y=675)
costos.winfo_class()

confirmacion = ttk.Button(text='Confirmar costos', command=confirmar)
confirmacion.place(x=1130,y=675)

boton_borrar = ttk.Button(text='Borrar datos', command=borrar_datos)
boton_borrar.place(x=1235,y=675)

root.mainloop()